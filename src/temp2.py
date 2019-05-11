# -*- coding: utf-8 -*-
import sip
import sys
import math
import yaml
import os
import uuid
import parser


from PyQt5.QtCore           import *
from PyQt5.QtGui            import *
from PyQt5.QtChart          import *
from PyQt5.QtWidgets        import * 
from swtw_icons.src.swtw_icons              import SWTW_GetIcon, SWTW_GetPixmap
from openpyxl               import Workbook
from copy                   import deepcopy
from functools              import partial
from pprint                 import pprint
from time                   import gmtime
from time                   import strftime
from datetime               import datetime
from datetime               import timedelta
from dateutil.relativedelta import relativedelta
from swtw_widgets.generic_widgets import *
from swtw_icons             import *
from ptcc_help.ptcc_help    import PTCC_HELP

_WDG_BACK_COLOR    = "#FFFFFF"
_WDG_BORDER_COLOR  = "#838487"
_WDG_BORDER_WIDTH  = "1"
_WDG_BORDER_RADIUS = "2"

PTCC_VERSION = "1.9.3"

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_StatusBar(QWidget):

    def __init__(self,app,parent):

        QWidget.__init__(self,parent)

        self.app          = app
        self.wdg_txt      = PTCC_WDG_Label("                                         ")
        self.wdg_progress = QProgressBar(self)
        self.wdg_progress.setFixedHeight(8)
        self.wdg_progress.setTextVisible(False)
        self.wdg_progress.setMinimum(0)
        self.wdg_progress.setMaximum(100)

        _font = QFont()
        _font.setPointSize(8)
        self.wdg_txt.setFont(_font)

        self.setFixedHeight(40)
        self.setFixedHeight(40)

        self.ly  = QHBoxLayout()
        self.ly.addWidget(self.wdg_txt) 
        self.ly.addWidget(self.wdg_progress)
        self.setLayout(self.ly)

        self.setStyleSheet("border: 0px solid white; margin: 0px;")

        self.wdg_txt.setStyleSheet("border: 1px solid gray; border-radius: 4px;")
        self.wdg_progress.setStyleSheet("border: 1px solid gray; border-radius: 4px;")

        self.stop()

    def start(self,withload=False):

        self.setVisible(True)
        self.msg("")
        self.progress(0)        
        self.wdg_progress.setVisible(withload)
        self.app.processEvents()

    def stop(self):

        self.msg("")
        self.progress(0)
        self.setVisible(False)

    def progress(self,value):

        self.wdg_progress.setValue(value)
        self.wdg_progress.update()
        self.app.processEvents()

    def msg(self,msg):

        self.wdg_txt.setText(msg)
        self.app.processEvents()

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Fields(QListWidget):

    def __init__(self,tooltip=""):

        QListWidget.__init__(self)
        self.setStyleSheet("background-color: white;")
        self.fields  = []
        self.tooltip = tooltip

    def register_callback_double_click(self,clbk):

        self.itemDoubleClicked.connect(clbk)

    def populate(self,fields):

        for _field in fields:

            self.addField(_field)

    def addField(self,field):

        if field not in self.fields:

            self.fields.append(field)

            _field_item = QListWidgetItem(field)

            _field_item.setToolTip(self.tooltip)

            self.addItem(_field_item)

    def removeItem(self,item):

        self.fields.remove(item.text())

        sip.delete(item)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Selection(QComboBox):

    def __init__(self):

        QComboBox.__init__(self)

        _css  = """  background-color: %s;
                     color: #000000;  
                     border: %spx solid gray;
                     border-color: %s;
                     border-radius: 3px;
                     font-size: 9pt;
                     font-family: Arial;

                     """ % (_WDG_BACK_COLOR,_WDG_BORDER_WIDTH,_WDG_BORDER_COLOR)


        self.setStyleSheet(_css)
        self.editTextChanged.connect(self.textChangedHandler)
        self.setEditable(True)
        self.data = []
        self._model_items = []


    def __create_completer(self):

        self.completer = QCompleter()

        self.completer.setFilterMode(Qt.MatchContains)

        self.setCompleter(self.completer)        

        self.model = QStringListModel()

        self.completer.setModel(self.model)        

        self._model_items = [list(_item.keys())[0] for _item in self.data]

        self.model.setStringList(self._model_items)

        self.completer.setModelSorting(QCompleter.CaseInsensitivelySortedModel)
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.completer.setCompletionRole(Qt.DisplayRole)

    def populate(self,data):

        self.data = data

        self.clear()
        for _item in self.data:

            self.addItem(list(_item.keys())[0])

        self.__create_completer()

        self.setCurrentIndex(0)

    def get_item_data(self,text):

        _data = ""

        for _index in range(len(self.data)):

            if list(self.data[_index].keys())[0] == text:

                _data = self.data[_index][list(self.data[_index].keys())[0]]
        
        return _data

    def textChangedHandler(self, text):
        if self._model_items:
            if str(text) in self._model_items or not len(text):
                self.setStyleSheet("background-color: white; border: %spx solid gray;" % _WDG_BORDER_WIDTH)
            else: 
                self.setStyleSheet("background-color: #ffcccc; border: %spx solid gray;" % _WDG_BORDER_WIDTH)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Small_Button(QPushButton):

    def __init__(self,icon_normal,icon_hover,clbk, tooltip=None):

        QPushButton.__init__(self)    

        _css  = """
            border: 0px solid gray;
            background-color: #FFFFFF;
        """ 

        self.setStyleSheet(_css)

        self.icon_normal = icon_normal
        self.icon_hover  = icon_hover
        self.clicked.connect(clbk)

        self.setIcon(QIcon(self.icon_normal))

        self.setToolTip(tooltip)

       
    def enterEvent(self, event):
        self.setIcon(QIcon(self.icon_hover))
    

    def leaveEvent(self, event):
        self.setIcon(QIcon(self.icon_normal))

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Button(QPushButton):

    def __init__(self,text,icon_normal,icon_hover,background):

        QPushButton.__init__(self,text)
        self.icon_normal = icon_normal
        self.icon_hover  = icon_hover
        self.background  = background

        self.setStyleSheet("background-color: %s" % (self.background,))
        self.setIcon(QIcon(self.icon_normal))
        
    def enterEvent(self, event):
        if self.icon_hover:
            self.setIcon(QIcon(self.icon_hover))
            self.setStyleSheet("background-color: %s" % (self.background,))      

    def leaveEvent(self, event):
        if self.icon_hover:
            self.setIcon(QIcon(self.icon_normal))
            self.setStyleSheet("background-color: %s" % (self.background,))

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_TextEdit(QLineEdit):

    def __init__(self,label):

        QLineEdit.__init__(self)

        self.setStyleSheet("background-color: white")

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Loading(QWidget):

    def __init__(self, parent = None):  

        QWidget.__init__(self, parent)

        palette = QPalette(self.palette())
        palette.setColor(palette.Background, Qt.transparent)
        self.setPalette(palette)

        self.title = PTCC_WDG_Label("")
        self.title.setAlignment(Qt.AlignHCenter)

        #Layout that contains text
        layout = QVBoxLayout(self)
        layout.addWidget(self.title)

    def paintEvent(self, event):

        painter = QPainter()
        painter.begin(self)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.fillRect(event.rect(), QBrush(QColor(255, 255, 255, 127)))
        painter.setPen(QPen(Qt.NoPen))
        
        for i in range(6):  
            if (self.counter / 5) % 6 == i:
                painter.setBrush(QBrush(QColor(20,20,20,240)))      
            else:
                painter.setBrush(QBrush(QColor(127, 127, 127)))     
            painter.drawEllipse(
                self.width()/2 + 15 * math.cos(2 * math.pi * i / 6.0) - 10,
                self.height()/2 + 15 * math.sin(2 * math.pi * i / 6.0) - 10, 10, 10)
            
        painter.end()      
        
        #move text 
        self.title.move(self.title.x(), self.height()/2 + self.height()/20)         
        self.update()                                                               

    def showEvent(self, event):
       
        self.timer = self.startTimer(100)
        self.counter = 0         

    def hideEvent(self, event):

        self.title.clear()      

    def timerEvent(self, event):
        
        self.counter += 1
        self.update()

    def setTextDisplay(self, value=''):

        self.title.setText(str(value))

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_PopUp(QMessageBox):

    def __init__(self,title,txt,msgtype="information"):

        QMessageBox.__init__(self)

        self.txt     = txt
        self.title   = title
        self.msgtype = msgtype

        self.draw_gui()

        self.exec_()

    def draw_gui(self):

        self.setWindowTitle(self.title)
        self.setWindowIcon(SWTW_GetIcon('5a5160191638c3bd554f883f6d94b8e25ffc5e61'))
        self.resize(400,40)
        self.setFixedSize(400,40)
        self.setStyleSheet("background-color: #ffffff")
        self.setText(self.txt)
        self.setWindowModality(Qt.ApplicationModal)

        if self.msgtype == "question":
            self.setIcon(QMessageBox.Question)
        else:
            if self.msgtype == "information":
                self.setIcon(QMessageBox.Information)
            else:
                if self.msgtype == "warning":
                    self.setIcon(QMessageBox.Warning)
                else:
                    if self.msgtype == "critical":
                        self.setIcon(QMessageBox.Critical)
                    else:
                        self.setIcon(QMessageBox.NoIcon)

        self.setStandardButtons(QMessageBox.Ok)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Question(QMessageBox):

    def __init__(self,title,txt,bt1,bt2):

        QMessageBox.__init__(self)

        self.txt        = txt
        self.title      = title
        self.bt1        = bt1
        self.bt2        = bt2
        self.status     = False
        self.is_pressed = False

        self.draw_gui()

        self.buttonClicked.connect(self.clbk_button_clicked)

        self.exec_()

    def draw_gui(self):

        self.setWindowTitle(self.title)
        self.setWindowIcon(SWTW_GetIcon('5a5160191638c3bd554f883f6d94b8e25ffc5e61'))
        self.resize(400,40)
        self.setFixedSize(400,40)
        self.setStyleSheet("background-color: #ffffff")
        self.setText(self.txt)
        self.setWindowModality(Qt.ApplicationModal)
       
        self.setIcon(QMessageBox.Question)

        self.addButton(self.bt1,QMessageBox.AcceptRole)
        self.addButton(self.bt2,QMessageBox.RejectRole)

    def clbk_button_clicked(self,button):

        self.is_pressed = True

        if button.text() == self.bt1:
            self.status = True
        else:
            self.status = False

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Credentials(QFrame):

    def __init__(self,server,port):

        QFrame.__init__(self)

        self.server = server
        self.port   = port

        self.draw_gui()

    def draw_gui(self):

        #windows properties
        self.setWindowTitle("Enter Credentials")
        self.setWindowIcon(SWTW_GetIcon('5a5160191638c3bd554f883f6d94b8e25ffc5e61'))
        self.resize(245,135)
        self.setFixedSize(245,135)
        self.setStyleSheet("background-color: #E5EBF7")
        self.setWindowModality(Qt.ApplicationModal)

        #main text
        self.lbl_txt_1  = PTCC_WDG_Label("Please type your user name and password for")
        self.lbl_txt_1.setWordWrap(true);
        self.lbl_txt_2  = PTCC_WDG_Label("%s:%s" % (self.server,self.port))


        #inputs
        self.username = PTCC_WDG_TextEdit("")
        self.password = PTCC_WDG_TextEdit("")
        self.username_lbl = PTCC_WDG_Label("User Name:")
        self.password_lbl = PTCC_WDG_Label("Password:  ")
        _username_layout  = QHBoxLayout()
        _password_layout  = QHBoxLayout()
        _username_layout.addWidget(self.username_lbl)
        _username_layout.addWidget(self.username)
        _password_layout.addWidget(self.password_lbl)
        _password_layout.addWidget(self.password)

        #buttons
        self.bt_ok     = QPushButton("OK")
        self.bt_cancel = QPushButton("Cancel")
        self.bt_help   = QPushButton("Help")

        #text layout vertical
        _text_layout  = QVBoxLayout()
        _text_layout.addWidget(self.lbl_txt_1)
        _text_layout.addWidget(self.lbl_txt_2)
        _text_layout.addLayout(_username_layout)
        _text_layout.addLayout(_password_layout)

        #button layout hotizontal
        _bt_layout  = QHBoxLayout()
        _bt_layout.addWidget(self.bt_ok)
        _bt_layout.addWidget(self.bt_cancel)
        _bt_layout.addWidget(self.bt_help)

        #global layout vertical
        _glb_layout  = QVBoxLayout()
        _glb_layout.addLayout(_text_layout)
        _glb_layout.addLayout(_bt_layout)

        self.setLayout(_glb_layout)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Tree(Widgets_QTreeWidget):

    def __init__(self,parent=None,usefind=False):

        QTreeWidget.__init__(self)

        _css  = ""
        _css += "background-color: %s;" % (_WDG_BACK_COLOR,)
        _css += "color: #000000;"  
        _css += "border: %spx solid gray;" % (_WDG_BORDER_WIDTH,)
        _css += "border-color: %s;" % (_WDG_BORDER_COLOR,)
        _css += "border-radius: %spx;" % (_WDG_BORDER_RADIUS,)
        _css += "selection-color: #000000;"
        _css += "selection-background-color: #c2d2ed;"
        _css += "font-family: Arial;"
        _css += "font-size: 9pt;"

        self.header_items = []
        self.usefind      = usefind
        self.parent       = parent 
        self.setStyleSheet(_css)
        self.setRootIsDecorated(True)
        self.setHeaderHidden(True)  

        self.menu_item = None


    def search_tree(self):

        self.search = PTCC_WDG_Tree_Find_Dialog( self)
        self.search.show()     

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Test_Tree(Widgets_QTreeWidget):

    def __init__(self,parent=None,usefind=False):

        QTreeWidget.__init__(self)

        _css  = ""
        _css += "background-color: %s;" % (_WDG_BACK_COLOR,)
        _css += "color: #000000;"  
        _css += "border: %spx solid gray;" % (_WDG_BORDER_WIDTH,)
        _css += "border-color: %s;" % (_WDG_BORDER_COLOR,)
        _css += "border-radius: %spx;" % (_WDG_BORDER_RADIUS,)
        _css += "selection-color: #000000;"
        _css += "selection-background-color: #c2d2ed;"
        _css += "font-family: Arial;"
        _css += "font-size: 9pt;"

        self.header_items = []
        self.usefind      = usefind
        self.parent       = parent 
        self.setStyleSheet(_css)
        self.setRootIsDecorated(True)
        self.setHeaderHidden(True)  

        self.menu_item = None

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Stack(QStackedWidget):

    def __init__(self):

        QStackedWidget.__init__(self)

        self.setStyleSheet("margin-right: 5px;")

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Rotated_HeaderView(QHeaderView):

    def __init__(self, parent=None):
        super(PTCC_WDG_Rotated_HeaderView, self).__init__(Qt.Horizontal, parent)
        self.setMinimumSectionSize(20)
        self.setSectionsClickable(True)

    def paintSection(self, painter, rect, logicalIndex ):
        painter.save()
        # translate the painter such that rotate will rotate around the correct point
        painter.translate(rect.x()+rect.width(), rect.y())
        painter.rotate(90)
        # and have parent code paint at this location
        newrect = QRect(0,0,rect.height(),rect.width())
        super(PTCC_WDG_Rotated_HeaderView, self).paintSection(painter, newrect, logicalIndex)
        painter.restore()

    def minimumSizeHint(self):
        size = super(PTCC_WDG_Rotated_HeaderView, self).minimumSizeHint()
        size.transpose()
        return size

    def sectionSizeFromContents(self, logicalIndex):
        size = super(PTCC_WDG_Rotated_HeaderView, self).sectionSizeFromContents(logicalIndex)
        size.transpose()
        return size

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Table(QTableWidget):

    def __init__(self,with_metadata=False,vertical_headers=True,with_large_v_headers=True,safe_label=True,horiz_sort=False,with_find=False):

        QTableWidget.__init__(self)

        _css_table = """

                    QWidget {
                        background-color: #ffffff;
                        color: #000000;
                        spacing: 0px;
                        padding: 0px;
                    }

                    QHeaderView::section {
                        background-color: #cecccc;
                        padding: 0px;
                        border: 1px solid #707070;
                        font-size: 7pt;
                        border-radius: 3px;
                        font-family: Arial;
                        color: #1e1e1e
                    }

                    QTableWidget {
                        gridline-color: #d3d3d3;
                        font-size: 9pt;
                        font-family: Arial;
                    }

                    QTableWidget QTableCornerButton::section {
                        background-color: #ffffff;
                        border: 1px solid #fffff8;
                    }
                    
                    QScrollBar:horizontal {
                     border: 1px solid #222222;
                     background: black;
                     height: 7px;
                     margin: 0px 16px 0 16px;
                    }
                    
                    QScrollBar::handle:horizontal
                    {
                          background: white;
                          min-height: 20px;
                          border-radius: 2px;
                    }
                    
                    QScrollBar::add-line:horizontal {
                          border: 1px solid #1b1b19;
                          border-radius: 2px;
                          background: QLinearGradient( x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #adaba9, stop: 1 #e0dfde);
                          width: 14px;
                          subcontrol-position: right;
                          subcontrol-origin: margin;
                    }
                    
                    QScrollBar::sub-line:horizontal {
                          border: 1px solid #1b1b19;
                          border-radius: 2px;
                          background: QLinearGradient( x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #adaba9, stop: 1 #e0dfde);
                          width: 14px;
                         subcontrol-position: left;
                         subcontrol-origin: margin;
                    }
                    
                    QScrollBar::right-arrow:horizontal, QScrollBar::left-arrow:horizontal
                    {
                          border: 1px solid black;
                          width: 1px;
                          height: 1px;
                          background: white;
                    }
                    
                    QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal
                    {
                          background: none;
                    }
                    
                    QScrollBar:vertical
                    {
                          background: black;
                          width: 7px;
                          margin: 16px 0 16px 0;
                          border: 1px solid #222222;
                    }
                    
                    QScrollBar::handle:vertical
                    {
                          background: white;
                          min-height: 20px;
                          border-radius: 2px;
                    }
                    
                    QScrollBar::add-line:vertical
                    {
                          border: 1px solid #1b1b19;
                          border-radius: 2px;
                          background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #adaba9, stop: 1 #e0dfde);
                          height: 14px;
                          subcontrol-position: bottom;
                          subcontrol-origin: margin;
                    }
                    
                    QScrollBar::sub-line:vertical
                    {
                          border: 1px solid #1b1b19;
                          border-radius: 2px;
                          background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #e0dfde, stop: 1 #adaba9);
                          height: 14px;
                          subcontrol-position: top;
                          subcontrol-origin: margin;
                    }
                    
                    QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical
                    {
                          border: 1px solid black;
                          width: 1px;
                          height: 1px;
                          background: white;
                    }
        """

        self.with_metadata = with_metadata
        self.safe_label    = safe_label
        self.horiz_sort    = horiz_sort
        self.with_find     = with_find

        self.setStyleSheet(_css_table)

        if vertical_headers:
            self.setHorizontalHeader(PTCC_WDG_Rotated_HeaderView(self))
        

        self.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        if with_large_v_headers:
            self.verticalHeader().setMinimumWidth(350)
            self.verticalHeader().setMaximumWidth(350)

        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        self.old_mouse_release  = self.mouseReleaseEvent
        self.mouseReleaseEvent  = self.table_mouse_release

    def __menu_view(self):

        pass

    def table_mouse_release(self, event):

        if event.button() == Qt.RightButton:

            self.popMenu = QMenu()
            self.popMenu.addAction(SWTW_GetIcon('8a480dfd6b6c153858f51b3923874aed0bf8414a'),"View in Separate Window", self.__menu_view)
            self.popMenu.exec_(self.popMenu.mapToGlobal(QCursor.pos()))

        self.old_mouse_release(event)

    def get_safe_label(self,label,size):

        _safe_label = label

        if len(label) > size:

            _safe_label = label[:(size - 5)] + "..."

        return _safe_label

    def add_horizontal_header(self,header_items):

        _headers = []
     
        self.setColumnCount(len(header_items))

        for _index in range(len(header_items)):

            if self.with_metadata:
                _label = header_items[_index][0]
            else:
                _label = header_items[_index]

            if self.safe_label:
                _header    = QTableWidgetItem(self.get_safe_label(_label,15))  
                _header.setToolTip(_label)
            else:
                _header    = QTableWidgetItem(_label)

            _headers

            if self.with_metadata:          
                _serial_cfg = header_items[_index][1]
                _header.setData(Qt.UserRole, _serial_cfg)

            self.setHorizontalHeaderItem(_index, _header)

        if self.horiz_sort:
            self.horizontalHeader().sectionClicked.connect(self.sort)

    def add_vertical_header(self,header_items):

        self.setRowCount(len(header_items))

        for _index in range(len(header_items)):

            if self.with_metadata:
                _label = header_items[_index][0]
            else:
                _label = header_items[_index]

            if self.safe_label:
                _header = QTableWidgetItem(self.get_safe_label(_label,70))
                _header.setToolTip(_label)
            else:
                _header = QTableWidgetItem(_label)
            
            if self.with_metadata:
                _serial_cfg = header_items[_index][1]
                _header.setData(Qt.UserRole, _serial_cfg)

            self.setVerticalHeaderItem(_index, _header)

    def add_items(self,items):

        for _row_index in range(len(items)):

            for _column_index in range(len(items[_row_index])):

                _item = QTableWidgetItem(str(items[_row_index][_column_index]))

                _item.setTextAlignment(Qt.AlignRight)

                _item.setFlags(_item.flags() ^ Qt.ItemIsEditable)

                self.setItem(_row_index, _column_index, _item)

    def add_group_items(self,items):

        for _row_index in range(len(items)):

            for _column_index in range(len(items[_row_index])):

                _group = items[_row_index][_column_index]

                if isinstance(_group,PTCC_Items_Group):

                    if _group.empty:
                        _value = "-"
                    else:
                        _value = str(len(_group))
                else:
                    _value = _group

                _item = QTableWidgetItem(_value)

                _item.setTextAlignment(Qt.AlignRight)

                _item.setFlags(_item.flags() ^ Qt.ItemIsEditable)

                self.setItem(_row_index, _column_index, _item)

                _item.setData(Qt.UserRole, _group)

    def get_user_data(self,item):

        if self.with_metadata:
            _serial_data = item.data(Qt.UserRole)
        else:
            _serial_data = None

        return _serial_data

    def sort(self,index):

        self.sortItems(index,Qt.DescendingOrder)

    def keyPressEvent(self, event):

        if self.with_find:

            # ctrl+F pressed
            _find_shortcut = (event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_F)  
            
            if _find_shortcut:
                self.search_table()

        super(PTCC_WDG_Table, self).keyPressEvent(event) 

    def search_table(self):

        self.search = PTCC_WDG_Table_Find_Dialog(self)
        self.search.show() 

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Tab(QTabWidget):

    def __init__(self):

        QTabWidget.__init__(self)

        _css  = """
        QTabWidget::pane { 
                            border: 1px solid #707070;
                            border-radius: 3px;
                         }

        QTabBar::tab {
                    border: 1px solid #707070;
                    border-bottom-color: #bfc0c1;
                    min-width: 55ex;
                    font-family: Arial;
                    padding: 2px;
                    border-radius: 3px;
                    color: #494949
                }

        QTabBar::tab:selected {
            border-color: gray;
            border-bottom-color: #707070; 
            background-color: #c1c1c1;
            border-radius: 3px;

        }

        QTabBar::tab:!selected {
            margin-top: 4px; /* make non-selected tabs look smaller */
        }

        """ 


        self.setStyleSheet(_css)

    def add_tab(self,label):

        _widget = QWidget()

        self.addTab(_widget,label)

        return _widget

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Label(QLabel):

    def __init__(self,txt=""):

        QLabel.__init__(self,txt)

        _css  = """
        background-color: %s;
        color: #000000; 
        border: 0px solid gray;
        font-family: Arial;

        """ % (_WDG_BACK_COLOR,)

        self.setStyleSheet(_css)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Splash(QSplashScreen):

    def __init__(self,app,parent):

        self.__width  = 299
        self.__height = 39
        self.parent  = parent


        QSplashScreen.__init__(self,parent,pixmap=SWTW_GetPixmap('038654c51da3c73c97a62188efd8886021469ed1'))

        _flags = self.windowFlags() | Qt.WindowStaysOnTopHint
        self.setWindowFlags(_flags)

        self.app          = app
        self.wdg_main     = QWidget(self)
        self.wdg_main.setFixedWidth(self.__width)
        self.wdg_main.setFixedHeight(self.__height)
        self.wdg_txt      = PTCC_WDG_Label("")
        self.wdg_txt.setFixedWidth(self.__width)
        self.wdg_progress = QProgressBar(self)
        self.wdg_progress.setTextVisible(False)
        self.wdg_progress.setFixedWidth(self.__width)
        self.wdg_progress.setFixedHeight(10)
        self.wdg_progress.setMinimum(0)
        self.wdg_progress.setMaximum(100)

        self.ly  = QVBoxLayout()
        self.ly.addWidget(self.wdg_txt)
        self.ly.addWidget(self.wdg_progress)


        self.wdg_main.setLayout(self.ly)

    def center(self):
        
        _geometry = self.parent.geometry()
        _x = (_geometry.x() + _geometry.width()) / 2
        _y = (_geometry.y() + _geometry.height()) / 2
        self.move(_x,_y)

    def start(self,withload=False):

        self.show()
        self.app.processEvents()
        self.wdg_progress.setVisible(withload)

    def stop(self):

        self.finish(self)

    def progress(self,value):

        self.wdg_progress.setValue(value)
        self.wdg_progress.update()

    def msg(self,msg):

        self.wdg_txt.setText(msg)
        self.app.processEvents()

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_PieSlice(QPieSlice):

    def __init__(self,label,items):

        QPieSlice.__init__(self,label,len(items))

        self.items = items

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Plot_Donut(object):

    def __init__(self,parent,color_rules=[]):

        self.chart_view  = QChartView()
        self.height      = None
        self.width       = None
        self.series      = None
        self.parent      = parent
        self.color_rules = color_rules

        self.chart_view.mouseReleaseEvent = self.chart_mouse_release

    def get_image_path(self):

        _path = os.path.join(
                                os.environ["HOMEPATH"], 
                                "Documents",
                                "ptcc_reports",
                                strftime("ptcc_po_snapshot_%d_%m_%Y_%H_%M_%S.png", gmtime()))

        return _path

    def __menu_save(self):

        _path = self.get_image_path()

        _image = self.chart_view.grab()

        _image.save(_path,"PNG")

        os.system(_path)

    def chart_mouse_release(self, event):

        if event.button() == Qt.RightButton:

            self.popMenu = QMenu()
            self.popMenu.addAction(SWTW_GetIcon('f4841f48907bdc6426ec80971922ca72f37efc5a'),"Take Snapshot", self.__menu_save)
            self.popMenu.exec_(self.popMenu.mapToGlobal(QCursor.pos()))

    def set_size_max(self,height,width):

        self.height = height
        self.width  = width

    def draw(self,title,labels,colors,values):
        if self.series:
            self.chart_view.chart().removeSeries(self.series)
        
        _labels = []

        for _label in labels:
            if _label != "":
                _labels.append(_label)
            else:
                _labels.append("None")

        self.series     = QPieSeries()

        _total = 0
        for _index in range(len(values)):
            _total += len(values[_index])

        for _index in range(len(values)):

            _value_raw  = len(values[_index])
            if _total > 0 and len(values[_index])> 0:
                _value_proc = round((100.0/float(_total))*float(len(values[_index])),1)
            else:
                _value_proc = 0

            _slice = PTCC_WDG_PieSlice("%s (%s) (%s %%)" % (_labels[_index],_value_raw,_value_proc),values[_index])
            _slice.setLabelColor(Qt.black)
            _slice.setLabelPosition(QPieSlice.LabelOutside)
            _slice.doubleClicked.connect(partial(self.show_query,slice=_slice))
            if colors != None:
                _slice.setColor(QColor(colors[_index]))
            _slice.hovered[bool].connect(partial(self.explode_slice, slice=_slice))
            self.series.append(_slice)

        self.chart_view.setRenderHint(QPainter.Antialiasing)
        self.chart = self.chart_view.chart()
        self.chart.legend().setVisible(True)
        self.chart.legend().setAlignment(Qt.AlignRight)
        self.chart.legend().setShowToolTips(True)
        self.chart.legend().setMarkerShape(QLegend.MarkerShapeCircle)

        self.chart.setTheme(QChart.ChartThemeLight)

        #self.chart.setTitle(title)
        self.chart.setAnimationOptions(QChart.AllAnimations)
        self.chart_view.chart().addSeries(self.series)
        if self.height and self.width:
            self.chart_view.setMaximumSize(self.height, self.width)

    def explode_slice(self, exploded, slice):

        if exploded:

            _start_angle = slice.startAngle()

            _end_angle   = slice.startAngle() + slice.angleSpan()

        else:
            _donut = slice.series()

            _donut.setPieStartAngle(0)

            _donut.setPieEndAngle(360)

        #slice.setLabelVisible(exploded)

        slice.setExploded(exploded)

        _the_marker = self.__get_slice_marker(slice)

        for _marker in self.chart.legend().markers():
            self.__set_marker_normal(_marker)

        if _the_marker:
            self.__set_marker_large(_the_marker)

    def __set_marker_large(self,marker):

            _font = QFont()
            _font.setPointSize(10)
            _font.setBold(True)
            marker.setFont(_font)

    def __set_marker_normal(self,marker):
        
            _font = QFont()
            _font.setPointSize(8)
            _font.setBold(False)
            marker.setFont(_font)

    def __get_slice_marker(self,slice): 

        _the_marker = None

        for _marker in self.chart.legend().markers():

            if slice.label() == _marker.label():
                _the_marker = _marker

        return _the_marker

    def show_query(self,slice):

        slice.items.show(self.parent,self.color_rules)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Plot_MultipleDonut(object):

    def __init__(self,parent,color_rules=[]):

        self.chart_view  = QChartView()
        self.height      = None
        self.width       = None
        self.series      = None
        self.donuts      = []
        self.parent      = parent
        self.color_rules = color_rules

        self.chart_view.mouseReleaseEvent = self.chart_mouse_release

    def get_image_path(self):

        _path = os.path.join(
                                os.environ["HOMEPATH"], 
                                "Desktop",
                                "ptcc_reports",
                                strftime("ptcc_po_snapshot_%d_%m_%Y_%H_%M_%S.png", gmtime()))

        return _path

    def __menu_save(self):

        _path = self.get_image_path()

        _image = self.chart_view.grab()

        _image.save(_path,"PNG")

        os.system(_path)

    def chart_mouse_release(self, event):

        if event.button() == Qt.RightButton:

            self.popMenu = QMenu()
            self.popMenu.addAction(SWTW_GetIcon('f4841f48907bdc6426ec80971922ca72f37efc5a'),"Take Snapshot", self.__menu_save)
            self.popMenu.exec_(self.popMenu.mapToGlobal(QCursor.pos()))

    def set_size_max(self,height,width):

        self.height = height
        self.width  = width

    def draw(self,title,labels,values,colors=None):
        self.donuts = []

        if self.series != None:
            for _series in self.series:
                self.chart_view.chart().removeSeries(_series)
            self.series = []

        self.series = []

        _labels       = []
        _nr_of_donuts = len(labels)
        _min_size     = 0.1
        _max_size     = 0.9

        for _label_group in labels:
            _new_label_group = []
            for _label in _label_group:
                if _label != "":
                    _new_label_group.append(_label)
                else:
                    _new_label_group.append("None")
            _labels.append(_new_label_group)

        for _donut_index in range(_nr_of_donuts):

            _series = QPieSeries()

            _total = 0
            for _index in range(len(values)):
                _total += len(values[_donut_index][_index])

            for _value_index in range(len(values[_donut_index])):

                _value_raw  = len(values[_donut_index][_value_index])
                if _total > 0 and len(values[_donut_index][_value_index])> 0:
                    _value_proc = round((100.0/float(_total))*float(len(values[_donut_index][_value_index])),1)
                else:
                    _value_proc = 0

                _slice = PTCC_WDG_PieSlice("%s (%s) (%s %%)" % (
                                                                _labels[_donut_index][_value_index],
                                                                _value_raw,
                                                                _value_proc),
                                                                values[_donut_index][_value_index])

                _slice.doubleClicked.connect(partial(self.show_query,slice=_slice))
                _slice.setLabelColor(Qt.black)
                if colors != None:
                    _slice.setColor(QColor(colors[_donut_index][_value_index]))

                _slice.setLabelPosition(QPieSlice.LabelOutside)
                _slice.hovered[bool].connect(partial(self.explode_slice, slice=_slice))
                _series.append(_slice)
                _series.setHoleSize(_min_size + _donut_index * (_max_size - _min_size) / _nr_of_donuts)
                _series.setPieSize(_min_size + (_donut_index + 1) * (_max_size - _min_size) / _nr_of_donuts)

            self.donuts.append(_series)
            self.chart_view.chart().addSeries(_series)
            self.series.append(_series)
        
        self.chart_view.setRenderHint(QPainter.Antialiasing)
        self.chart = self.chart_view.chart()
        self.chart.legend().setVisible(True)
        self.chart.legend().setAlignment(Qt.AlignRight)
        self.chart.legend().setMarkerShape(QLegend.MarkerShapeCircle)
        self.chart.setTheme(QChart.ChartThemeLight)
        #self.chart.setTitle(title)
        self.chart.setAnimationOptions(QChart.AllAnimations)
        
        if self.height and self.width:
            self.chart_view.setMaximumSize(self.height, self.width)

    def explode_slice(self, exploded, slice):

        if exploded:

            _slice_start_angle = slice.startAngle()
            _slice_end_angle   = slice.startAngle() + slice.angleSpan()

            _donut = slice.series()
            _series_index = self.donuts.index(_donut)
            for _index in range(_series_index + 1, len(self.donuts)):
                self.donuts[_index].setPieStartAngle(_slice_end_angle)
                self.donuts[_index].setPieEndAngle(360 + _slice_start_angle)
        else:
            for _donut in self.donuts:
                _donut.setPieStartAngle(0)
                _donut.setPieEndAngle(360)

        #slice.setLabelVisible(exploded)

        slice.setExploded(exploded)

        _the_marker = self.__get_slice_marker(slice)

        for _marker in self.chart.legend().markers():
            self.__set_marker_normal(_marker)

        if _the_marker:
            self.__set_marker_large(_the_marker)

    def __set_marker_large(self,marker):

            _font = QFont()
            _font.setPointSize(10)
            _font.setBold(True)
            marker.setFont(_font)

    def __set_marker_normal(self,marker):
        
            _font = QFont()
            _font.setPointSize(8)
            _font.setBold(False)
            marker.setFont(_font)

    def __get_slice_marker(self,slice): 

        _the_marker = None

        for _marker in self.chart.legend().markers():

            if slice.label() == _marker.label():
                _the_marker = _marker

        return _the_marker

    def show_query(self,slice):

        slice.items.show(self.parent,self.color_rules)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Plot_Calendar(QDateEdit):

    def __init__(self,min_date,crt_date):

        QDateEdit.__init__(self)

        self.min_year  = min_date[0]
        self.min_month = min_date[1]
        self.min_day   = min_date[2]

        self.crt_year  = crt_date[0]
        self.crt_month = crt_date[1]
        self.crt_day   = crt_date[2]

        self.draw()

    def draw(self):


        _css = """ QDateEdit {
                    border: 2px solid lightgray;
                    border-radius: 5px;
                }
                 
                QDateEdit::drop-down {
                    subcontrol-origin: padding;
                    subcontrol-position: top right;
                    width: 15px;
                 
                    border-left-width: 1px;
                    border-left-color: darkgray;
                    border-left-style: solid;
                    border-top-right-radius: 3px;
                    border-bottom-right-radius: 3px;
                }
                 
                QDateEdit::down-arrow {
                    image: url(:/DateEditTest/Resources/Combobox Arrow.png);
                }
                 
                QDateEdit::drop-down:hover {
                    /* Does this work with QDateEdit??? */
                    background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #d8d7d7, stop: 1 #e9e9e9);
                }
        """

        self.setStyleSheet(_css)

        self.setCalendarPopup(True)
        self.setDisplayFormat("dd/MM/yyyy")
        self.setMinimumDate(QDate(self.min_year,self.min_month,self.min_day))
        self.setDate(QDate(self.crt_year,self.crt_month,self.crt_day))

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Plot_Lines_Periods(QObject):

    draw_sgn  = pyqtSignal(list) 
    clear_sgn = pyqtSignal()

    def __init__(self):

        QObject.__init__(self)

        self.chart          = QChart()
        self.chart_view     = QChartView(self.chart)
        self.timer          = QTimer(self)
        self.lines          = []
        self.visible_lines  = []
        self._y_labels_vis  = [True]
        self.series         = []

        self.timer.timeout.connect(self.__register_markers)

        self.draw_sgn.connect(self.__draw_values)
        self.clear_sgn.connect(self.clear)

        self.old_mouse_release = self.chart_view.mouseReleaseEvent
        self.chart_view.mouseReleaseEvent = self.chart_mouse_release

    def get_image_path(self):

        _path = os.path.join(
                                os.environ["HOMEPATH"], 
                                "Desktop",
                                "ptcc_reports",
                                strftime("ptcc_po_snapshot_%d_%m_%Y_%H_%M_%S.png", gmtime()))

        return _path

    def __menu_save(self):

        _path = self.get_image_path()

        _image = self.chart_view.grab()

        _image.save(_path,"PNG")

        os.system(_path)

    def chart_mouse_release(self, event):

        if event.button() == Qt.RightButton:

            self.popMenu = QMenu()
            self.popMenu.addAction(SWTW_GetIcon('f4841f48907bdc6426ec80971922ca72f37efc5a'),"Take Snapshot", self.__menu_save)
            self.popMenu.exec_(self.popMenu.mapToGlobal(QCursor.pos()))

        self.old_mouse_release(event)

    def clear(self):

        for _serie in self.series:
            self.chart.removeSeries(_serie)

        self.series = []

    def create_lines(self,values,y_labels,y_label_colors):

        _nr_of_lines  = len(y_labels)

        for _line_index in range(_nr_of_lines):

            _label  = y_labels[_line_index]
            _values = []
            _color  = y_label_colors[_line_index]

            for _value_index in range(len(values[_line_index])):

                _values.append([_value_index,values[_line_index][_value_index]])

            self.lines.append([_label,_values,_color])

    def draw(self,title,values,y_labels,y_label_colors,x_labels):

        self.visible_lines = []
        self.lines         = []
        self.x_labels      = x_labels
        self.title         = title
        self._y_labels_vis = [True] * len(y_labels)
        self.y_labels      = y_labels

        self.create_lines(values,y_labels,y_label_colors)

        self.visible_lines = deepcopy(self.lines)

        self.__draw_values(self.visible_lines)

    def __draw_values(self,lines):

        _x_axis = QBarCategoryAxis()
        _x_axis.append(self.x_labels)
        
        for _line in lines:

            _serie = QLineSeries()

            for _value in _line[1]:
                _point = self.__get_point(_value)
                _serie.append(self.__get_point(_value))

            _serie.setPointsVisible(True)
            _serie.setColor(QColor(_line[2]))

            if _line[0] == "":
                _serie.setName("None")
            else:
                _serie.setName(_line[0])

            self.chart.addSeries(_serie)
            self.series.append(_serie)

        _x_axis.setLabelsAngle(-90)
        self.chart.createDefaultAxes()
        self.chart.setAxisX(_x_axis,_serie)        

        self.chart.setTitle(self.title)
        self.chart.setAnimationOptions(QChart.SeriesAnimations)
        self.chart.setAnimationDuration(500)
        self.chart.legend().setVisible(True)
        self.chart.legend().setAlignment(Qt.AlignRight)
        self.chart_view.setRenderHint(QPainter.Antialiasing)

        self.__set_markers_state()

        self.timer.setSingleShot(True)
        self.timer.start(600)

    def __get_point(self,value):

        _x = value[0]
        _y = value[1]

        return QPointF(_x, _y)

    def __register_markers(self):

        for _marker in self.chart.legend().markers():
            _marker.clicked.connect(partial(self.__clbk_marker_click,_marker))
            _marker.hovered.connect(self.__clbk_marker_hovered)

    def __set_markers_state(self):

        for _marker in self.chart.legend().markers():

            _the_label = _marker.label() 
            if _the_label == "None":
                _the_label = ""

            _the_index = self.y_labels.index(_the_label)
            if self._y_labels_vis[_the_index]:
                _alpha = 1.0
            else:
                _alpha = 0.3

            _label_brush        = _marker.labelBrush()
            _label_brush_color  = _label_brush.color()
            _label_brush_color.setAlphaF(_alpha)
            _label_brush.setColor(_label_brush_color)
            _marker.setLabelBrush(_label_brush)

            _marker_brush       = _marker.brush()
            _marker_brush_color = _marker_brush.color()
            _marker_brush_color.setAlphaF(_alpha)
            _marker_brush.setColor(_marker_brush_color)
            _marker.setBrush(_marker_brush)

            _marker_pen         = _marker.pen()
            _marker_pen_color   = _marker_pen.color()
            _marker_pen_color.setAlphaF(_alpha)
            _marker_pen.setColor(_marker_pen_color)
            _marker.setPen(_marker_pen)

    def __clbk_marker_click(self,marker):

        _the_label = marker.label() 

        if _the_label == "None":
            _the_label = ""

        _the_index = self.y_labels.index(_the_label)
        self._y_labels_vis[_the_index] = not self._y_labels_vis[_the_index]

        self.visible_lines = deepcopy(self.lines)

        for _index in range(len(self.visible_lines)):

            if not self._y_labels_vis[_index]:

                self.visible_lines[_index][1] = [[0,0]]

        self.clear_sgn.emit()
        self.draw_sgn.emit(self.visible_lines)

    def __clbk_marker_hovered(self, status):

        if status:
            QApplication.setOverrideCursor(Qt.PointingHandCursor)
        else:
            QApplication.setOverrideCursor(Qt.ArrowCursor)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Plot_StackedBar_Periods(QObject):

    draw_sgn = pyqtSignal(list) 
    clear_sgn = pyqtSignal()

    def __init__(self):

        QObject.__init__(self)

        self.chart            = QChart()
        self.chart_view       = QChartView(self.chart)    
        self.series           = QStackedBarSeries()
        self.timer            = QTimer(self)
        self._title           = "" 
        self._values          = [[0]] 
        self._y_labels        = [""] 
        self._y_label_colors  = ["#000000"] 
        self._x_labels        = [""] 
        self._y_labels_vis    = [True]
        self.has_values       = False

        self.timer.timeout.connect(self.__register_markers)

        self.__draw_values(self._values)

        self.draw_sgn.connect(self.__draw_values)
        self.clear_sgn.connect(self.clear)

        self.old_mouse_release = self.chart_view.mouseReleaseEvent
        self.chart_view.mouseReleaseEvent = self.chart_mouse_release

    def get_image_path(self):

        _path = os.path.join(
                                os.environ["HOMEPATH"], 
                                "Desktop",
                                "ptcc_reports",
                                strftime("ptcc_po_snapshot_%d_%m_%Y_%H_%M_%S.png", gmtime()))

        return _path

    def __menu_save(self):

        _path = self.get_image_path()

        _image = self.chart_view.grab()

        _image.save(_path,"PNG")

        os.system(_path)

    def chart_mouse_release(self, event):

        if event.button() == Qt.RightButton:

            self.popMenu = QMenu()
            self.popMenu.addAction(SWTW_GetIcon('f4841f48907bdc6426ec80971922ca72f37efc5a'),"Take Snapshot", self.__menu_save)
            self.popMenu.exec_(self.popMenu.mapToGlobal(QCursor.pos()))

        self.old_mouse_release(event)

    def clear(self):

        self.chart.removeSeries(self.series)

    def draw(self,title,values,y_labels,y_label_colors,x_labels):

        self._title           = title
        self._values          = values
        self._y_labels        = y_labels
        self._y_label_colors  = y_label_colors
        self._x_labels        = x_labels
        self._y_labels_vis    = [True] * len(y_labels)

        self.__draw_values(self._values)

    def __draw_values(self,values):

        self.has_values = True

        _y_labels = []
        for _label in self._y_labels:
            if "" != _label:
                _y_labels.append(_label)
            else:
                _y_labels.append("None")

        self.series  = QStackedBarSeries()
        _x_axis      = QBarCategoryAxis()
        _barsets     = []

        for _index in range(len(_y_labels)):

            _barset = QBarSet(_y_labels[_index])
            _barset.setColor(QColor(self._y_label_colors[_index]))
            _barsets.append(_barset)

            for _value in values[_index]:
                _barsets[-1].append(_value)

        for _barset in _barsets:
            self.series.append(_barset)

        self.chart.addSeries(self.series)
        self.chart.setTitle(self._title)        
        _x_axis.append(self._x_labels)
        _x_axis.setLabelsAngle(-90)
        self.chart.createDefaultAxes()
        self.chart.setAxisX(_x_axis, self.series)
        self.chart.setAnimationOptions(QChart.SeriesAnimations)
        self.chart.setAnimationDuration(500)
        self.chart.legend().setVisible(True)
        self.chart.legend().setAlignment(Qt.AlignRight)
        self.chart_view.setRenderHint(QPainter.Antialiasing)

        self.__set_markers_state()

        self.timer.setSingleShot(True)
        self.timer.start(600)

        self.__register_hover() 

    def __register_hover(self):

        self.series.hovered.connect(self.__clbk_bar_hover)

    def __register_markers(self):

        for _marker in self.chart.legend().markers():

            _marker.clicked.connect(partial(self.__clbk_marker_click,_marker))
            _marker.hovered.connect(self.__clbk_marker_hovered)

    def __set_markers_state(self):

        for _marker in self.chart.legend().markers():

            _the_label = _marker.label() 

            if _the_label == "None":
                _the_label = ""

            _the_index = self._y_labels.index(_the_label)
            if self._y_labels_vis[_the_index]:
                _alpha = 1.0
            else:
                _alpha = 0.3

            _label_brush        = _marker.labelBrush()
            _label_brush_color  = _label_brush.color()
            _label_brush_color.setAlphaF(_alpha)
            _label_brush.setColor(_label_brush_color)
            _marker.setLabelBrush(_label_brush)

            _marker_brush       = _marker.brush()
            _marker_brush_color = _marker_brush.color()
            _marker_brush_color.setAlphaF(_alpha)
            _marker_brush.setColor(_marker_brush_color)
            _marker.setBrush(_marker_brush)

            _marker_pen         = _marker.pen()
            _marker_pen_color   = _marker_pen.color()
            _marker_pen_color.setAlphaF(_alpha)
            _marker_pen.setColor(_marker_pen_color)
            _marker.setPen(_marker_pen)

    def __clbk_bar_hover(self,status,index,barset):

        if status:

            _x_label = self._x_labels[index]
            _y_label = barset.label()
            _value   = barset.at(index)

            _point = self.chart_view.mapFromGlobal(QCursor.pos());

            _x = _point.x()
            _y = _point.y()

            self.hover_rect = QGraphicsRectItem(self.chart)
            self.hover_rect.setBrush(QBrush(QColor("#f9ec9f")))
            self.hover_rect.setRect(_x, _y, 120, 60)  

            self.hover_txt  = QGraphicsTextItem (self.hover_rect)  
            self.hover_txt.setPos(_x, _y,)    
            self.hover_txt.setPlainText("%s\n%s\n%s" % (_y_label,_value,_x_label))
            self.hover_rect.setZValue(10)

        self.hover_rect.setVisible(status)

    def __clbk_marker_click(self,marker):

        _the_label = marker.label() 

        if _the_label == "None":
            _the_label = ""

        _the_index = self._y_labels.index(_the_label)
        self._y_labels_vis[_the_index] = not self._y_labels_vis[_the_index]

        _values         = []

        for _index in range(len(self._y_labels)):
            if self._y_labels_vis[_index]:
                _values.append(self._values[_index])
            else:
                _values.append([0] * len(self._values[_index]))

        self.clear_sgn.emit()
        self.draw_sgn.emit(_values)

    def __clbk_marker_hovered(self, status):

        if status:
            QApplication.setOverrideCursor(Qt.PointingHandCursor)
        else:
            QApplication.setOverrideCursor(Qt.ArrowCursor)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Plot_Bar_Periods(QObject):

    draw_sgn = pyqtSignal(list) 
    clear_sgn = pyqtSignal()

    def __init__(self):

        QObject.__init__(self)

        self.chart            = QChart()
        self.chart_view       = QChartView(self.chart)    
        self.series           = QBarSeries()
        self.timer            = QTimer(self)
        self._title           = "" 
        self._values          = [[0]] 
        self._y_labels        = [""] 
        self._y_label_colors  = ["#000000"] 
        self._x_labels        = [""] 
        self._y_labels_vis    = [True]
        self.has_values       = False

        self.timer.timeout.connect(self.__register_markers)

        self.__draw_values(self._values)

        self.draw_sgn.connect(self.__draw_values)
        self.clear_sgn.connect(self.clear)

        self.old_mouse_release = self.chart_view.mouseReleaseEvent
        self.chart_view.mouseReleaseEvent = self.chart_mouse_release

    def get_image_path(self):

        _path = os.path.join(
                                os.environ["HOMEPATH"], 
                                "Desktop",
                                "ptcc_reports",
                                strftime("ptcc_po_snapshot_%d_%m_%Y_%H_%M_%S.png", gmtime()))

        return _path

    def __menu_ptc(self):

        pass

    def __menu_save(self):

        _path = self.get_image_path()

        _image = self.chart_view.grab()

        _image.save(_path,"PNG")

        os.system(_path)

    def chart_mouse_release(self, event):

        if event.button() == Qt.RightButton:

            self.popMenu = QMenu()
            self.popMenu.addAction(SWTW_GetIcon('f4841f48907bdc6426ec80971922ca72f37efc5a'),"Take Snapshot", self.__menu_save)
            self.popMenu.exec_(self.popMenu.mapToGlobal(QCursor.pos()))

        self.old_mouse_release(event)

    def clear(self):

        self.chart.removeSeries(self.series)

    def draw(self,title,values,y_labels,y_label_colors,x_labels):

        self._title           = title
        self._values          = values
        self._y_labels        = y_labels
        self._y_label_colors  = y_label_colors
        self._x_labels        = x_labels
        self._y_labels_vis    = [True] * len(y_labels)

        self.__draw_values(self._values)
       
    def __draw_values(self,values):

        self.has_values = True

        _y_labels = []
        for _label in self._y_labels:
            if "" != _label:
                _y_labels.append(_label)
            else:
                _y_labels.append("None")

        self.series  = QBarSeries()
        _x_axis      = QBarCategoryAxis()
        _barsets     = []

        for _index in range(len(_y_labels)):

            _barset = QBarSet(_y_labels[_index])
            _barset.setColor(QColor(self._y_label_colors[_index]))
            _barsets.append(_barset)

            for _value in values[_index]:
                _barsets[-1].append(_value)

        for _barset in _barsets:
            self.series.append(_barset)

        self.chart.addSeries(self.series)
        self.chart.setTitle(self._title)        
        _x_axis.append(self._x_labels)
        _x_axis.setLabelsAngle(-90)
        self.chart.createDefaultAxes()
        self.chart.setAxisX(_x_axis, self.series)
        self.chart.setAnimationOptions(QChart.SeriesAnimations)
        self.chart.setAnimationDuration(500)
        self.chart.legend().setVisible(True)
        self.chart.legend().setAlignment(Qt.AlignRight)
        self.chart_view.setRenderHint(QPainter.Antialiasing)

        self.__set_markers_state()

        self.timer.setSingleShot(True)
        self.timer.start(600)
            
        self.__register_hover() 

    def __register_hover(self):

        self.series.hovered.connect(self.__clbk_bar_hover)

    def __register_markers(self):

        for _marker in self.chart.legend().markers():

            _marker.clicked.connect(partial(self.__clbk_marker_click,_marker))
            _marker.hovered.connect(self.__clbk_marker_hovered)

    def __set_markers_state(self):

        for _marker in self.chart.legend().markers():

            _the_label = _marker.label() 
            if _the_label == "None":
                _the_label = ""

            _the_index = self._y_labels.index(_the_label)
            if self._y_labels_vis[_the_index]:
                _alpha = 1.0
            else:
                _alpha = 0.3

            _label_brush        = _marker.labelBrush()
            _label_brush_color  = _label_brush.color()
            _label_brush_color.setAlphaF(_alpha)
            _label_brush.setColor(_label_brush_color)
            _marker.setLabelBrush(_label_brush)

            _marker_brush       = _marker.brush()
            _marker_brush_color = _marker_brush.color()
            _marker_brush_color.setAlphaF(_alpha)
            _marker_brush.setColor(_marker_brush_color)
            _marker.setBrush(_marker_brush)

            _marker_pen         = _marker.pen()
            _marker_pen_color   = _marker_pen.color()
            _marker_pen_color.setAlphaF(_alpha)
            _marker_pen.setColor(_marker_pen_color)
            _marker.setPen(_marker_pen)

    def __clbk_bar_hover(self,status,index,barset):

        if status:

            _x_label = self._x_labels[index]
            _y_label = barset.label()
            _value   = barset.at(index)

            _point = self.chart_view.mapFromGlobal(QCursor.pos());

            _x = _point.x()
            _y = _point.y()

            self.hover_rect = QGraphicsRectItem(self.chart)
            self.hover_rect.setBrush(QBrush(QColor("#f9ec9f")))
            self.hover_rect.setRect(_x, _y, 100, 60)  

            self.hover_txt  = QGraphicsTextItem (self.hover_rect)  
            self.hover_txt.setPos(_x, _y,)    
            self.hover_txt.setPlainText("%s\n%s\n%s" % (_y_label,_value,_x_label))
            self.hover_rect.setZValue(10)

        self.hover_rect.setVisible(status)

    def __clbk_marker_click(self,marker):

        _the_label = marker.label() 

        if _the_label == "None":
            _the_label = ""

        _the_index = self._y_labels.index(_the_label)
        self._y_labels_vis[_the_index] = not self._y_labels_vis[_the_index]

        _values         = []

        for _index in range(len(self._y_labels)):
            if self._y_labels_vis[_index]:
                _values.append(self._values[_index])
            else:
                _values.append([0] * len(self._values[_index]))

        self.clear_sgn.emit()
        self.draw_sgn.emit(_values)

    def __clbk_marker_hovered(self, status):
        if status:
            QApplication.setOverrideCursor(Qt.PointingHandCursor)
        else:
            QApplication.setOverrideCursor(Qt.ArrowCursor)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_RadioButton(QRadioButton):

    def __init__(self,label):

        QRadioButton.__init__(self,label)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_CheckBox(QCheckBox):

    def __init__(self,label):

        QCheckBox.__init__(self,label)

        _css = """
        font-family: Arial

        """

        self.setStyleSheet(_css)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Tree_Find_Dialog(QWidget):


    def __init__(self,tree):

        QWidget.__init__(self,tree)

        self._items          = []
        self._item_idx       = 0
        self.tree            = tree
        self.selected_column = 0
           
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Popup )         
            
        #write line
        self.line = QLineEdit()
        self.line.setContextMenuPolicy(Qt.NoContextMenu)    
        self.line.setStyleSheet('border: 2px solid gray;')
        self.line.returnPressed.connect(self._find_next) 
        self.line.textEdited.connect(self._find)

        #find previous button
        self._find_prev_button = QToolButton()
        self._find_prev_button.setStyleSheet("background: transparent; border-radius: 0px")
        self._find_prev_button.setIcon(SWTW_GetIcon('8964290011c997b1b3b55e8c84b72745c7671a9a'))
        self._find_prev_button.clicked.connect(self._find_prev)                 

        #find next button
        self._find_next_button = QToolButton()
        self._find_next_button.setStyleSheet("background: transparent; border-radius: 0px")
        self._find_next_button.setIcon(SWTW_GetIcon('7d692e927b39bdd591599cd93dad6cdb89d0a724'))
        self._find_next_button.clicked.connect(self._find_next)           

        #close button
        self._close_button = QToolButton()
        self._close_button.setStyleSheet("background: transparent; border-radius: 0px")
        self._close_button.setIcon(SWTW_GetIcon('c4c80912bacc504d1b0259c9f2fe36c548b5aca0'))
        self._close_button.clicked.connect(self._close)

        #column selection
        self._selection = PTCC_WDG_Selection()
        self._selection.setFixedWidth(150)
        self._selection.currentIndexChanged.connect(self.selection_change)
        _data = []
        for _index in range(len(self.tree.header_items)):
            _data.append({self.tree.header_items[_index]:_index})
        self._selection.populate(_data)

        box = QHBoxLayout(self)
        box.addWidget(self.line)
        box.addWidget(self._selection)
        box.addWidget(self._find_next_button)
        box.addWidget(self._find_prev_button)        
        box.addWidget(self._close_button)
        box.setSizeConstraint(3)
        box.setContentsMargins(8, 8, 5, 5)
        self.setLayout(box)
        
        QShortcut(QKeySequence(Qt.Key_Escape), self).activated.connect(self._close)           
        QShortcut(QKeySequence(Qt.CTRL + Qt.Key_F), self).activated.connect(self._close)    

        #Animation
        self._animation = QPropertyAnimation(self,b"geometry")
        self._animation.setTargetObject(self)
        self._animation.setDuration(150)
        self._animation.setEasingCurve(QEasingCurve.Linear)           

    def selection_change(self):

        self.selected_column = self._selection.currentText()

        self.selected_column = self._selection.get_item_data(self.selected_column)

    def showEvent(self, event): 

        _point = self.tree.mapToGlobal(QPoint(self.tree.geometry().x(),self.tree.geometry().y()))

        _x,_y = _point.x(),_point.y()
        
        self._animation.setStartValue( QRect(_x, _y + 200, self.width(), self.height()) )
        self._animation.setEndValue( QRect(_x, _y + 30, self.width(), self.height()) )
        self._animation.start() 
        
        self.line.setFocus()

    def _close(self):

        _point = self.tree.mapToGlobal(QPoint(self.tree.geometry().x(),self.tree.geometry().y()))

        _x,_y = _point.x(),_point.y()

        self._animation.setStartValue(QRect(_x, _y + 30, self.width(), self.height()))
        self._animation.setEndValue(QRect(_x, _y + 200, self.width(), self.height()))
        self._animation.start()
        self._animation.finished.connect(self.close)

    def _find(self):

        self._items     = self.tree.findItems(self.word, Qt.MatchContains | Qt.MatchRecursive, self.selected_column)
        self._item_idx  = 0

    def _find_next(self):

        try:            
            _item           = self._items[self._item_idx]   
            self._item_idx -= 1
        except:
            if self._items == []:
                #no matches found
                _item = QTreeWidgetItem()
            else:
                #end of matches
                self._item_idx = 0
                _item = self._items[self._item_idx] 

        self.tree.scrollToItem(_item)

        if _item.parent():

            _item.parent().setExpanded(True)        

        self.tree.setCurrentItem(_item,self.selected_column,QItemSelectionModel.ClearAndSelect)   
        self.tree.setFocus()   

    def _find_prev(self):
        
        try:    
            self._item_idx += 1
            _item = self._items[self._item_idx]     
        except:
            if self._items == []:
                #no matches found
                _item = QTreeWidgetItem()
            else:
                #end of matches
                self._item_idx  = len(self._items)-1        
                _item = self._items[self._item_idx] 
            
        self.tree.scrollToItem(_item)
        self.tree.setCurrentItem(_item,self.selected_column,QItemSelectionModel.ClearAndSelect)

    @property
    def word(self):

        return str(self.line.text())

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Table_Find_Dialog(QWidget):

    def __init__(self,table):

        QWidget.__init__(self,table)

        self._items          = []
        self._item_idx       = 0
        self.table           = table
           
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Popup )         
            
        #write line
        self.line = QLineEdit()
        self.line.setContextMenuPolicy(Qt.NoContextMenu)    
        self.line.setStyleSheet('border: 2px solid gray;')
        self.line.returnPressed.connect(self._find_next) 
        self.line.textEdited.connect(self._find)

        #find previous button
        self._find_prev_button = QToolButton()
        self._find_prev_button.setStyleSheet("background: transparent; border-radius: 0px")
        self._find_prev_button.setIcon(SWTW_GetIcon('8964290011c997b1b3b55e8c84b72745c7671a9a'))
        self._find_prev_button.clicked.connect(self._find_prev)                 

        #find next button
        self._find_next_button = QToolButton()
        self._find_next_button.setStyleSheet("background: transparent; border-radius: 0px")
        self._find_next_button.setIcon(SWTW_GetIcon('7d692e927b39bdd591599cd93dad6cdb89d0a724'))
        self._find_next_button.clicked.connect(self._find_next)           

        #close button
        self._close_button = QToolButton()
        self._close_button.setStyleSheet("background: transparent; border-radius: 0px")
        self._close_button.setIcon(SWTW_GetIcon('c4c80912bacc504d1b0259c9f2fe36c548b5aca0'))
        self._close_button.clicked.connect(self._close)

        box = QHBoxLayout(self)
        box.addWidget(self.line)
        box.addWidget(self._find_next_button)
        box.addWidget(self._find_prev_button)        
        box.addWidget(self._close_button)
        box.setSizeConstraint(3)
        box.setContentsMargins(8, 8, 5, 5)
        self.setLayout(box)
        
        QShortcut(QKeySequence(Qt.Key_Escape), self).activated.connect(self._close)           
        QShortcut(QKeySequence(Qt.CTRL + Qt.Key_F), self).activated.connect(self._close)    

        #Animation
        self._animation = QPropertyAnimation(self,b"geometry")
        self._animation.setTargetObject(self)
        self._animation.setDuration(150)
        self._animation.setEasingCurve(QEasingCurve.Linear)          

    def showEvent(self, event): 

        _point = self.table.mapToGlobal(QPoint(self.table.geometry().x(),self.table.geometry().y()))

        _x,_y = _point.x(),_point.y()
        
        self._animation.setStartValue( QRect(_x + 30, _y + 30,     self.width(), self.height()) )
        self._animation.setEndValue(   QRect(_x, _y,     self.width(), self.height()) )
        self._animation.start() 
        
        self.line.setFocus()

    def _close(self):

        _point = self.table.mapToGlobal(QPoint(self.table.geometry().x(),self.table.geometry().y()))

        _x,_y = _point.x(),_point.y()

        self._animation.setStartValue(QRect(_x, _y, self.width(), self.height()))
        self._animation.setEndValue(QRect(_x + 30, _y + 30, self.width(), self.height()))
        self._animation.start()
        self._animation.finished.connect(self.close)

    def _find(self):

        self._items     = self.table.findItems(self.word, Qt.MatchContains | Qt.MatchRecursive)
        self._item_idx  = 0

    def _find_next(self):

        try:            
            _item           = self._items[self._item_idx]   
            self._item_idx -= 1
        except:
            if self._items == []:
                #no matches found
                _item = QTableWidgetItem()
            else:
                #end of matches
                self._item_idx = 0
                _item = self._items[self._item_idx] 

        self.table.scrollToItem(_item)      
        self.table.setCurrentItem(_item,QItemSelectionModel.ClearAndSelect)   
        self.table.setFocus()   

    def _find_prev(self):
        
        try:    
            self._item_idx += 1
            _item = self._items[self._item_idx]     
        except:
            if self._items == []:
                #no matches found
                _item = QTableWidgetItem()
            else:
                #end of matches
                self._item_idx  = len(self._items)-1        
                _item = self._items[self._item_idx] 
            
        self.table.scrollToItem(_item)
        self.table.setCurrentItem(_item,QItemSelectionModel.ClearAndSelect)
        self.table.setFocus()   

    @property
    def word(self):

        return str(self.line.text())

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Table_Query(QMainWindow):

    def __init__(self,parent,headers,values,color_rules=[]):

        QMainWindow.__init__(self,parent)

        self.headers     = headers
        self.values      = values
        self.parent      = parent
        self.color_rules = color_rules

        self.draw_gui()

        self.show()

    def draw_gui(self):

        self.setWindowTitle("PTCC Query")
        self.setWindowIcon(SWTW_GetIcon('5a5160191638c3bd554f883f6d94b8e25ffc5e61'))
        self.resize(800,800)
        self.setStyleSheet("background-color: #FFFFFF")
        #self.setWindowModality(Qt.ApplicationModal)

        self.draw_table()

        self.ctrl_wdg = QWidget()

        self.excel_bt = PTCC_WDG_Small_Button(SWTW_GetIcon('68359959371beb83fcd68fd11770af72a16d55d0'),SWTW_GetIcon('37a3fa6917686adcd0254f905c8dd50a89e824d1'),self.excel, "Export as Excel")
        self.excel_bt.setFixedWidth(40)

        self.search_bt = PTCC_WDG_Small_Button(SWTW_GetIcon('f12404a4b24f4ee746b13893bb7d7e9e67dafd97'),SWTW_GetIcon('b96c3d64ea61d5e592580aff0617a17918a7dd9e'),self.search_table, "Search Table")
        self.search_bt.setFixedWidth(40)

        self.toolbar = QToolBar()
        self.toolbar.addWidget(self.excel_bt)
        self.toolbar.addWidget(self.search_bt)


        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.toolbar)
        self.main_layout.addWidget(self.table)

        self.ctrl_wdg.setLayout(self.main_layout)

        self.setCentralWidget(self.ctrl_wdg)

    def draw_table(self):

        self.table = QTableWidget()
        self.table.setStyleSheet(self.get_table_css())
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionsMovable(True)
        self.add_table_horizontal_header(self.headers)
        self.add_table_items(self.values)
        self.table.cellDoubleClicked.connect(self.open_item)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSortingEnabled(True)


    def get_table_css(self):
        _css = """
                    QWidget {
                        background-color: #ffffff;
                        color: #000000;
                        spacing: 0px;
                        padding: 0px;
                    }

                    QHeaderView::section {
                        background-color: #ffffff;
                        padding: 0px;
                        border: 1px solid #707070;
                        font-size: 7pt;
                        border-radius: 3px;
                        font-family: Arial;
                        color: #1e1e1e
                    }

                    QTableWidget {
                        gridline-color: #d3d3d3;
                        font-size: 9pt;
                        font-family: Arial;
                    }

                    QTableWidget QTableCornerButton::section {
                        background-color: #ffffff;
                        border: 1px solid #fffff8;
                    }
                     QScrollBar:horizontal {
                     border: 1px solid #222222;
                     background: black;
                     height: 7px;
                     margin: 0px 16px 0 16px;
                    }
                    
                    QScrollBar::handle:horizontal
                    {
                          background: white;
                          min-height: 20px;
                          border-radius: 2px;
                    }
                    
                    QScrollBar::add-line:horizontal {
                          border: 1px solid #1b1b19;
                          border-radius: 2px;
                          background: QLinearGradient( x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #adaba9, stop: 1 #e0dfde);
                          width: 14px;
                          subcontrol-position: right;
                          subcontrol-origin: margin;
                    }
                    
                    QScrollBar::sub-line:horizontal {
                          border: 1px solid #1b1b19;
                          border-radius: 2px;
                          background: QLinearGradient( x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #adaba9, stop: 1 #e0dfde);
                          width: 14px;
                         subcontrol-position: left;
                         subcontrol-origin: margin;
                    }
                    
                    QScrollBar::right-arrow:horizontal, QScrollBar::left-arrow:horizontal
                    {
                          border: 1px solid black;
                          width: 1px;
                          height: 1px;
                          background: white;
                    }
                    
                    QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal
                    {
                          background: none;
                    }
                    
                    QScrollBar:vertical
                    {
                          background: black;
                          width: 7px;
                          margin: 16px 0 16px 0;
                          border: 1px solid #222222;
                    }
                    
                    QScrollBar::handle:vertical
                    {
                          background: white;
                          min-height: 20px;
                          border-radius: 2px;
                    }
                    
                    QScrollBar::add-line:vertical
                    {
                          border: 1px solid #1b1b19;
                          border-radius: 2px;
                          background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #adaba9, stop: 1 #e0dfde);
                          height: 14px;
                          subcontrol-position: bottom;
                          subcontrol-origin: margin;
                    }
                    
                    QScrollBar::sub-line:vertical
                    {
                          border: 1px solid #1b1b19;
                          border-radius: 2px;
                          background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #e0dfde, stop: 1 #adaba9);
                          height: 14px;
                          subcontrol-position: top;
                          subcontrol-origin: margin;
                    }
                    
                    QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical
                    {
                          border: 1px solid black;
                          width: 1px;
                          height: 1px;
                          background: white;
                    }

        """

        return _css

    def add_table_horizontal_header(self,header_items):

        self.headers = header_items
     
        self.table.setColumnCount(len(header_items))

        for _index in range(len(header_items)):

            _label      = header_items[_index][0]
            _header     = QTableWidgetItem(_label)  
            _user_data  = header_items[_index][1]
            _serial_cfg = yaml.dump(_user_data)
            _header.setData(Qt.UserRole, _serial_cfg)

            self.table.setHorizontalHeaderItem(_index, _header)

        self.table.horizontalHeader().sectionClicked.connect(self.sort)

    def get_cell_color(self,label):

        _color = "#FFFFFF"

        for _rule in self.color_rules:
            if label == _rule[0]:
                _color = _rule[1]

        return QColor(_color)

    def add_table_items(self,items):

        self.items = items

        self.table.setRowCount(len(items))

        for _row_index in range(len(items)):

            for _column_index in range(len(items[_row_index])):

                _item = QTableWidgetItem(str(items[_row_index][_column_index]))

                _cell_color = self.get_cell_color(str(items[_row_index][_column_index]))

                _item.setBackground(_cell_color)

                _item.setFlags(_item.flags() ^ Qt.ItemIsEditable)

                self.table.setItem(_row_index, _column_index, _item)

    def get_table_user_data(self,item):

        _serial_data = str(item.data(Qt.UserRole))
        _data =  yaml.load(_serial_data)

        return _data

    def open_item(self,row,column):

        if hasattr(self.parent.parent,"im"):

            self.parent.parent.im.open_item_in_gui(self.table.item(row,0).text())
        else:
            self.parent.parent.parent.im.open_item_in_gui(self.table.item(row,0).text())
    
    def get_excel_name(self):

        name = "ptcc_query"

        _ptcc_report_path = os.path.join(os.environ["HOMEPATH"],"Desktop","ptcc_reports")

        if not os.path.exists(_ptcc_report_path):

            os.mkdir(_ptcc_report_path)

        _path = os.path.join(
                                _ptcc_report_path,
                                "%s_%s" % ( name, strftime("%d_%m_%Y_%H_%M_%S.xlsx", gmtime()) )
                            )

        return _path

    def excel(self):

        _status_bar = None

        if hasattr(self.parent,"statusbar"):
            _status_bar = self.parent.statusbar
        else:
            if hasattr(self.parent,"parent"):
                if hasattr(self.parent.parent,"statusbar"):
                    _status_bar = self.parent.parent.statusbar
                else:
                    _status_bar = None
            else:
                _status_bar = None
                
        if _status_bar:
            _status_bar.start()
            _status_bar.msg("generating report") 

        _book = Workbook()
        _book.remove(_book.active)
        
        self.excel_report_table(_book)

        _path = self.get_excel_name()

        _book.save(_path)

        if _status_bar:
            _status_bar.stop()

        os.system(_path)

    def search_table(self):

        self.search = PTCC_WDG_Table_Find_Dialog(self.table)
        self.search.show() 

    def keyPressEvent(self, event):

        # ctrl+F pressed
        _find_shortcut = (event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_F)  
        
        if _find_shortcut:
            self.search_table()

        super(PTCC_WDG_Table_Query, self).keyPressEvent(event) 

    def excel_write_cell(self,sheet,row,column,data):

        _safe_data = str(data).encode("utf-8","ignore")

        sheet.cell(row=row,column=column).value = _safe_data

    def excel_write_row(self,sheet,row,column,data):

        for _value in data:

            self.excel_write_cell(sheet,row,column,_value)

            column += 1

    def excel_report_table(self,book):

        _sheet     = book.create_sheet(title="Query")
        _start_row    = 1
        _start_column = 1

        self.excel_write_row(   _sheet,    _start_row,     _start_column, [_header[0] for _header in self.headers])

        _row = _start_row + 1

        for _value in self.items:
            self.excel_write_row(_sheet, _row, _start_column, _value)
            _row += 1

    def sort(self,index):

        self.table.sortItems(index,Qt.AscendingOrder)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_CustomWDG_Button(QPushButton):
    def __init__(self,icon,icon_click,colour_background_click,fct_click=None):

        QPushButton.__init__(self)
        self.icon                     = icon
        self.icon_click               = icon_click
        self.colour_background_click  = colour_background_click
        self.fct_click                = fct_click
        self.setStyleSheet("border: 0px solid gray;")
        self.setFixedSize(QSize(20,20))
        self.setIcon(self.icon)
        self.setIconSize(QSize(20,20))
        self.pressed.connect(self.__pressed)
        self.released.connect(self.__release)
        
    def __pressed(self):
        self.setIcon(self.icon_click)
        self.fct_click()

    def __release(self):
        self.setIcon(self.icon)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_List_Selection(QWidget):

    def __init__(self):
        QWidget.__init__(self)
        self.add_button     = PTCC_WDG_Small_Button(SWTW_GetIcon('15c643e72a69b2fa4b701240035260b5de4c9ded'), SWTW_GetIcon('3af7e41c7e2a8fa03c6adb673d43fef5ee41c4c8'), self.__add_elements, "Add Elements")
        self.delete_button  = PTCC_WDG_Small_Button(SWTW_GetIcon('a97cd58e3326a9786136c661140d0c7f7f5837c7'), SWTW_GetIcon('e5e469325c8bcf05a0e4dbfdac423d314ee41ebf'), self.__remove_elements, "Remove Elements")
        self.up_button      = PTCC_WDG_Small_Button(SWTW_GetIcon('6c8cde2178f2c537283f70ed201cdb61417fcffb'), SWTW_GetIcon('c31ca1c25352f493bd728987613ae6df1c18ba6c'), self.__up, "Move Up")
        self.up_button.setMinimumHeight(25)
        self.down_button    = PTCC_WDG_Small_Button(SWTW_GetIcon('f213bba72ff9bc558c64e253a5e13406a98d172e'), SWTW_GetIcon('8fd19b8029c5fa9c63f9e14d3f70851de0e32a9f'), self.__down, "Move Down")
       
        self.tb             = QToolBar()
        self.tb.setOrientation(Qt.Vertical)
        self.tb.addWidget(self.add_button)
        self.tb.addWidget(self.delete_button)
        self.tb.addSeparator()
        self.tb.addWidget(self.up_button)
        self.tb.addWidget(self.down_button)


        
        self.list1          = QListWidget()
        self.list1.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.list2          = QListWidget()
        self.list2.setSelectionMode(QAbstractItemView.ExtendedSelection)

        self.main_layout    = QHBoxLayout()
        self.main_layout.addWidget(self.list2)
        self.main_layout.addWidget(self.tb)
        self.main_layout.addWidget(self.list1)

    
        self.setLayout(self.main_layout)
        

    def __add_elements(self):
        __new_elements   = self.__get_selected_items_text(self.list2.selectedItems())
        __elements_list1 = self.__get_all_items_text_ListWidget(self.list1)
        for __element in __new_elements:
            if not __element in __elements_list1:
               self.list1.addItem(__element)

    def __remove_elements(self):
        for __element in self.list1.selectedItems():
            self.list1.takeItem(self.list1.row(__element))

    def __up(self):
        __selected_items = self.list1.selectedItems() 
        if len(__selected_items) == 1:
            __selected_item         = __selected_items[0]
            __selected_index        = self.list1.indexFromItem(__selected_item).row()
            if __selected_index != 0:
                __previous_element      = self.list1.item(__selected_index-1)
                __previous_element_text = __previous_element.text()
                __previous_element.setText(__selected_item.text())
                __selected_item.setText(__previous_element_text)
                self.list1.setCurrentItem(__previous_element)

    def __down(self):
        __selected_items = self.list1.selectedItems() 
        if len(__selected_items) == 1:
            __selected_item         = __selected_items[0]
            __selected_index        = self.list1.indexFromItem(__selected_item).row()
            if __selected_index != self.list1.count()-1:
                __next_element          = self.list1.item(__selected_index+1)
                __next_element_text   = __next_element.text()
                __next_element.setText(__selected_item.text())
                __selected_item.setText(__next_element_text)
                self.list1.setCurrentItem(__next_element)

    def __get_all_items_text_ListWidget(self,listWidget):
        items = []
        for index in range(listWidget.count()):
            items.append(listWidget.item(index).text())
        return items

    def __get_selected_items_text(self,items_objects):
        return [item.text() for item in items_objects]
        
    def list1_elements(self):
        return __get_all_items_text_ListWidget(self.list1)

    def populate(self,list1,list2):
        self.list1.clear()
        self.list2.clear()
        if list1 != None:
            for item in list1:
                self.list1.addItem(item)
        if list2 != None:
            for item in list2:
                self.list2.addItem(item)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_Console(QTextEdit):
    
    def __init__(self):
        QTextEdit.__init__(self)
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setReadOnly(True)
        self.setTextColor(Qt.white)
        self.setStyleSheet(self.get_css())
        self.message_log = ""

    def get_css(self):
        return """  border: 1px solid gray;
                    padding: 0px 0px 12px 5px;
                    background-color: rgba( 0, 0, 0, 60% );
        """

    def setBounds(self, x, y, w, h):
        
        self.setGeometry(x,y, w, h)
        
    def setfixedsize(self, x, y):
        self.setFixedSize(x, y)
        
    def setBgcolor(self, color):
        
        self.setStyleSheet("background-color: "+color)
        
    def setMsg(self, message, color=None):
        msg = "[ "+str(datetime.now().strftime("%Y-%M-%d %H:%M:%S"))+"]\t" + message
        self.message_log += msg + "\n"

        if None != color:
            self.setTextColor(color)
        else:
            self.setTextColor(Qt.white)
        self.append(msg)
        QApplication.processEvents()
        
    def getLog(self):
        return self.message_log
        
    def clearLog(self):
        self.clear()
        self.message_log = ""

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Table_Checkbox(QTableWidget):

    def __init__(self):

        QTableWidget.__init__(self)

        self.horizontalHeader().hide()
        self.verticalHeader().hide()

        self.selected_items = []

        self.mouseReleaseEvent = self.mouse_release

        self.table_items = []

    def get_safe_label(self,label):

        _safe_label = label

        if len(label) > 15:

            _safe_label = label[:10] + "..."

        return _safe_label

    def add_items(self,items):

        self.setRowCount(len(items))
        self.setColumnCount(1)

        for _index in range(len(items)):

            _item = PTCC_WDG_CheckBox(str(items[_index]))
            _item.stateChanged.connect(partial(self.state_change,item=_item))

            self.table_items.append(_item)

            self.setCellWidget(_index, 0, _item)

    def state_change(self,state,item):
        if Qt.Checked == state:
            self.selected_items.append(item.text())
        else:
            if Qt.Unchecked == state:
                self.selected_items.remove(item.text())

    def get_selected_items(self):

        return self.selected_items

    def __select_all(self):

        for _item in self.table_items:

            _item.setCheckState(Qt.Checked)

    def __deselect_all(self):

        for _item in self.table_items:

            _item.setCheckState(Qt.Unchecked)

    def mouse_release(self,event):

        if event.button() == Qt.RightButton:

            self.popMenu = QMenu()
            self.popMenu.addAction(SWTW_GetIcon('836b4d076077202c00f2fbd10c605023bb2bbfe5'),"Select All", self.__select_all)
            self.popMenu.addAction(SWTW_GetIcon('b49bc191b2a8c689d3d25431dde459e769349b8f'),"Deselect All", self.__deselect_all)
            self.popMenu.exec_(self.popMenu.mapToGlobal(QCursor.pos()))

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_About(QDialog):
    
    def __init__(self,parent,name):

        QDialog.__init__(self)

        self.name = name

        self.draw_gui()

    def draw_gui(self):

        global PTCC_VERSION
       
        self.setStyleSheet("border: 0px solid gray; ")

        font = QFont()
        font.setPointSize(10)
        font.setBold(True)
        
        #self.setWindowFlags(Qt.FramelessWindowHint)
        
        self.setWindowIcon(SWTW_GetIcon('5a5160191638c3bd554f883f6d94b8e25ffc5e61'))

        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        _logo = PTCC_WDG_Label()
        _logo.setPixmap(SWTW_GetPixmap("5115bc7c7a0197b7c4b314058b34324467cad688"))
        _logo.setStyleSheet("background-color: #ffffff;border: 0px;")

        #About text
        _about_info = QTextEdit()
        _about_info.setAlignment(Qt.AlignCenter)    
        _about_info.setReadOnly(True)
        _about_info.setTextInteractionFlags(Qt.NoTextInteraction| Qt.TextSelectableByKeyboard)
        _about_info.setContextMenuPolicy(Qt.NoContextMenu)

        _about_info.append('\n\n\n')
        _about_info.append(self.name)
        _about_info.append('\nVersion: %s'% PTCC_VERSION)
  
        _about_info.setStyleSheet("background-color: #ffffff;border: 0px;")
        _about_info.setFont(font)
        _about_info.setReadOnly(True)

        #SPACER
        _spacer = QWidget()             
        _spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        self.setFixedSize(320,200)
        self.setWindowTitle('About PTCC PO')

        _layout = QGridLayout() 
        _layout.addWidget(_logo,       0,0)
        _layout.addWidget(_about_info, 0,1)  
        _layout.setContentsMargins(0, 0, 0, 0)
        _layout.setAlignment(Qt.AlignTop)
        self.setLayout(_layout) 

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_Items_Group(object):

    def __init__(self,items=None,unique_attr=None,empty=False):

        self.unique_attr = unique_attr
        self.empty       = empty

        self.objects = []

        if items != None:
            for _item in items:
                if self.__is_unique(_item):
                    self.objects.append(_item)
        else:
            self.objects = []

    def add(self,obj):

        _is_add = True

        if self.unique_attr != None:

            _is_add = self.__is_unique(obj)

        if _is_add:

            self.objects.append(obj)  

    def __is_unique(self,obj):

        _is_unique = True

        if self.unique_attr != None:

            _value     = getattr(obj,self.unique_attr)
            _values    = [getattr(_object,self.unique_attr) for _object in self.objects]
            _is_unique = _value not in _values

        return _is_unique

    def index(self,obj):

        return [_object.id for _object in self.objects].index(obj.id)

    def count(self,obj):

        return [_object.id for _object in self.objects].count(obj.id)

    def __iter__(self):

        for obj in self.objects:

            yield obj

    def __getitem__(self,index):

        return self.objects[index]

    def __len__(self):

        return len(self.objects)

    def __contains__(self,obj):

        return obj.id in [_object.id for _object in self.objects]

    def __add__(self,group):

        if self.unique_attr != None:

            for _object in group.objects:

                if self.__is_unique(_object):

                    self.objects.append(_object)
        else:
            self.objects += group.objects

        return self

    def __sub__(self,group):

        _diff = PTCC_Items_Group(unique_attr=self.unique_attr)

        for _object in self.objects:

            if _object.id not in [_gr_obj.id for _gr_obj in group.objects]:

                _diff.add(_object)

        return _diff

    def find_all(self,obj):

        _objects = []

        for _obj_item in self.objects:

            if _obj_item.id == obj.id:

                _objects.append(_obj_item)

        return _objects

    def find_by_attribute(self,attribute,value):

        _object = None

        for _obj in self.objects:

            if getattr(_obj,attribute) == value:

                _object = _obj

        return _object

    def show(self,parent,rules=[]):
        
        if len(self.objects) > 0:

            _headers = self.objects[0].as_header()
            _values  = []

            for _object in self.objects:

                _values.append(_object.as_table_item())

            _query = PTCC_WDG_Table_Query(parent,_headers,_values,rules)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Isolate_Table(QMainWindow):

    def __init__(self,parent,title,horizontal_headers,vertical_headers,values,color_rules,donut_colors):

        QMainWindow.__init__(self,parent)

        self.parent             = parent
        self.horizontal_headers = horizontal_headers
        self.vertical_headers   = vertical_headers
        self.donut_colors       = donut_colors
        self.values             = values
        self.color_rules        = color_rules
        self.pivot              = "Columns"
        self.pivot_ranges       = None
        self.title              = title

        if len(self.values) > 0:
            self.draw_gui()

            self.show()

    def draw_gui(self):

        self.setWindowTitle("PTCC Table %s" % (self.title,))
        self.setWindowIcon(SWTW_GetIcon('5a5160191638c3bd554f883f6d94b8e25ffc5e61'))
        self.resize(800,800)
        self.setStyleSheet("background-color: #FFFFFF")

        self.ctrl_wdg = QWidget()

        self.btn_columns     = PTCC_WDG_Small_Button(SWTW_GetIcon('6d6adcc1f4af6d886de0a646d982f137f4eaee03'),SWTW_GetIcon('6d6adcc1f4af6d886de0a646d982f137f4eaee03'),self.__cbk_btn_columns, "Columns Pivot")
        self.btn_rows        = PTCC_WDG_Small_Button(SWTW_GetIcon('934977ece9f10b498795c494c40bfc1f2835c4e9'),SWTW_GetIcon('af91033457b1d1547f52f6ab0642a1f0e2248cc4'),self.__cbk_btn_rows, "Rows Pivot")
        self.excel_report_bt = PTCC_WDG_Small_Button(SWTW_GetIcon("68359959371beb83fcd68fd11770af72a16d55d0"),SWTW_GetIcon('37a3fa6917686adcd0254f905c8dd50a89e824d1'),self.excel_report, "Export as Excel")

        self.excel_report_bt.setFixedWidth(40)
        self.btn_rows.setFixedWidth(40)
        self.btn_columns.setFixedWidth(40)

        self.toolbar = QToolBar()

        self.toolbar.addWidget(self.btn_columns)
        self.toolbar.addWidget(self.btn_rows)
        self.toolbar.addSeparator()
        self.toolbar.addWidget(self.excel_report_bt)

        self.table = PTCC_WDG_Table(with_metadata=True)
        self.table.cellDoubleClicked.connect(self.show_query)

        self.donut = PTCC_WDG_Plot_Donut(self,color_rules=self.color_rules)

        self.table_layout = QHBoxLayout()

        self.table_layout.addWidget(self.table)
        self.table_layout.addWidget(self.donut.chart_view)

        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.toolbar)
        self.main_layout.addLayout(self.table_layout)      

        self.ctrl_wdg.setLayout(self.main_layout)

        self.setCentralWidget(self.ctrl_wdg)

        self.populate_table()

        _donut_values = []

        if len(self.values) > 0:
            if len(self.values[0][-1]) > 0:
                
                for _value in self.values[0][:-1]:
                    _donut_values.append(_value)

                self.donut.draw(
                                "",
                                self.horizontal_headers[:-1],
                                self.donut_colors,
                                _donut_values)

    def add_metadata_to_columns(self,columns):

        _headers = []

        for _header_item in columns:
            _headers.append([_header_item,{"data":_header_item}])

        return _headers

    def populate_table(self):

        _vertical_headers = []

        for _header in self.vertical_headers:
            if _header == "":
                _vertical_headers.append("None")
            else:
                _vertical_headers.append(_header)
                
        _horizontal_headers = []

        for _header in self.horizontal_headers:
            if _header == "":
                _horizontal_headers.append("None")
            else:
                _horizontal_headers.append(_header)
                
        self.table.clear()
        self.table.add_horizontal_header(self.add_metadata_to_columns(_horizontal_headers))       
        self.table.add_vertical_header(_vertical_headers)
        self.table.add_group_items(self.values)

        self.table.itemSelectionChanged.connect(self.clbk_table_select)
    
    def clbk_table_select(self):

        _ranges = self.get_table_ranges(self.table)        

        self.pivot_ranges = self.get_table_range_values(self.table,_ranges)

        self.draw_pivot_data()

    def show_query(self,row,column):

        _item        = self.table.item(row,column)
        _items_group = self.table.get_user_data(_item)

        if isinstance(_items_group,PTCC_Items_Group):
            _items_group.show(self,self.color_rules)

    def get_table_ranges(self,table):

        _ranges_raw = table.selectedRanges()
        _ranges     = []   

        for _range_raw in _ranges_raw:

            _top    = _range_raw.topRow()
            _bottom = _range_raw.bottomRow()
            _left   = _range_raw.leftColumn()
            _right  = _range_raw.rightColumn()

            _column_labels = []
            for _column_index in range(_left,_right + 1,1):

                _take_data = True
                _label = table.horizontalHeaderItem(_column_index).text()

                if _label == "Total":
                    if _left != _right or len(_ranges_raw) > 1:
                        _take_data = False

                if _take_data:
                    _data = table.get_user_data(table.horizontalHeaderItem(_column_index))
                    _data = _data['data']
                    _column_labels.append([_column_index,_data])

            _row_labels = []
            for _row_index in range(_top,_bottom + 1,1):

                _take_data = True
                _label = table.verticalHeaderItem(_row_index).text()

                if _label == "Total":
                    if _top != _bottom or len(_ranges_raw) > 1:
                        _take_data = False

                if _take_data:
                    _data = table.get_user_data(table.verticalHeaderItem(_row_index))
                    if isinstance(_data,dict):
                        _data = _data['data']
                    _row_labels.append([_row_index,_data])

            _ranges.append([_column_labels,_row_labels])

        return _ranges

    def get_table_range_values(self,table,ranges):

        for _range in ranges:
            
            _column_range = _range[0]
            _row_range    = _range[1]

            for _column in _column_range:

                _values = []

                for _row in _row_range:
                    _item = table.item(_row[0],_column[0])
                    if _item:
                        _values.append(table.get_user_data(_item))

                _column.append(_values)

            for _row in _row_range:

                _values = []

                for _column in _column_range:
                    _item = table.item(_row[0],_column[0])
                    if _item:
                        _values.append(table.get_user_data(_item))

                _row.append(_values)

        return ranges

    def draw_pivot_data(self):

        if self.pivot_ranges:

            _data = self.calculate_table_range_values(self.pivot_ranges)

            if self.pivot == "Columns":

                self.draw_table_range_donut_columns(_data)

            if self.pivot == "Rows":

                self.draw_table_range_donut_rows(_data)

    def calculate_table_range_values(self,ranges):

        _data = {}

        for _range in ranges:

            if self.pivot == "Columns":

                _column_range = _range[0]

                for _column in _column_range:

                    _sum = PTCC_Items_Group()
                    for _group in _column[2]:
                        _sum += _group

                    if _column[1] not in list(_data.keys()):
                        _data.update({_column[1]:_sum})
                    else:
                        _data[_column[1]] += _sum
        
            if self.pivot == "Rows":

                _row_range = _range[1]

                for _row in _row_range:

                    _sum = PTCC_Items_Group()
                    for _group in _row[2]:
                        _sum += _group

                    if _row[1] not in list(_data.keys()):
                        _data.update({_row[1]:_sum})
                    else:
                        _data[_row[1]] += _sum

        if len(_data.keys()) > 1:
            if "Total" in list(_data.keys()):
                del _data["Total"]

        return _data

    def draw_table_range_donut_columns(self,data):

        _total        = 0
        _donut_values = []

        for _marker in data:
            _total += len(data[_marker])

        if _total != 0:
            for _marker in data:
                _donut_values.append( data[_marker])

            _markers = data.keys()
            _colors  = [] 

            for _marker in _markers:

                if _marker == "None":

                    _marker = ""

            _index = self.horizontal_headers.index(_marker)
            _colors.append(self.donut_colors[_index])

            self.donut.draw(
                            "",
                            _markers,
                            self.donut_colors,
                            _donut_values)

    def draw_table_range_donut_rows(self,data):

        _total        = 0
        _donut_values = []

        for _marker in data:
            _total += len(data[_marker])

        if _total != 0:
            for _marker in data:
                _donut_values.append(data[_marker])

            _markers = data.keys()
            _colors  = [] 

            self.donut.draw(
                            "",
                            _markers,
                            None,
                            _donut_values)

    def __cbk_btn_rows(self):

        self.pivot = "Rows"

        self.btn_rows.icon_normal = SWTW_GetIcon('af91033457b1d1547f52f6ab0642a1f0e2248cc4')
        self.btn_columns.icon_normal = SWTW_GetIcon('1c237cca2d3985ed4486848743ef1767a74e5c70')
        self.btn_columns.setIcon(SWTW_GetIcon('1c237cca2d3985ed4486848743ef1767a74e5c70'))

        self.draw_pivot_data()

    def __cbk_btn_columns(self):

        self.pivot = "Columns"

        self.btn_columns.icon_normal = SWTW_GetIcon('6d6adcc1f4af6d886de0a646d982f137f4eaee03')
        self.btn_rows.icon_normal = SWTW_GetIcon('934977ece9f10b498795c494c40bfc1f2835c4e9')
        self.btn_rows.setIcon(SWTW_GetIcon('934977ece9f10b498795c494c40bfc1f2835c4e9'))

        self.draw_pivot_data()

    def get_report_path(self,name):

        _ptcc_report_path = os.path.join(os.environ["HOMEPATH"],"Desktop","ptcc_reports")

        if not os.path.exists(_ptcc_report_path):

            os.mkdir(_ptcc_report_path)

        _path = os.path.join(
                                _ptcc_report_path,
                                "%s_%s" % ( name, strftime("%d_%m_%Y_%H_%M_%S.xlsx", gmtime()) )
                            )

        return _path

    def excel_report(self):

        _book = Workbook()
        _book.remove(_book.active)


        _sheet     = _book.create_sheet(title="Table")
        _start_row    = 1
        _start_column = 1

        self.excel_write_row(   _sheet, _start_row,     _start_column, [""] + self.horizontal_headers)
        self.excel_write_column(_sheet, _start_row + 1, _start_column, [_header[0] for _header in self.vertical_headers])

        _row = _start_row + 1

        for _value in self.values:
            self.excel_write_row(_sheet, _row, _start_column + 1, [len(_obj_value) for _obj_value in _value])
            _row += 1

        _path = self.get_report_path("ptcc_po_table")

        _book.save(_path)

        os.system(_path)

    def excel_write_cell(self,sheet,row,column,data):

        _safe_data = str(data).encode("utf-8","ignore")

        sheet.cell(row=row,column=column).value = _safe_data

    def excel_write_row(self,sheet,row,column,data):

        for _value in data:

            self.excel_write_cell(sheet,row,column,_value)

            column += 1

    def excel_write_column(self,sheet,row,column,data):

        for _value in data:

            self.excel_write_cell(sheet,row,column,_value)

            row += 1

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_Filter(object):

    def __init__(self,name,filter_dict):

        self.name        = name
        self.description = ""
        self.code        = ""
        self.filter_dict = filter_dict
        self.id          = ""

        self.load()

    def load(self):

        if "description" in self.filter_dict.keys():
            self.description = self.filter_dict["description"]

        if "code" in self.filter_dict.keys():
            self.code = self.filter_dict["code"]

        self.id = uuid.uuid4().hex

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Filter(QDialog):

    def __init__(self,parent,path,name):

        QDialog.__init__(self,parent)

        self.filters          = [] 
        self.path             = path
        self.name             = name
        self.code_visible     = False
        self.selected_filters = []
        self.operator         = "or"
        self.code             = ""
        self.filter           = None 
        self.parent           = parent
        self.version          = "0.0.0"

        self.draw_gui()

        _error = self.load()

    def display(self):

        self.populate()

        self.show()

    def populate(self):

        self.filter_tree.clear()

        for _filter_category in self.filters:

            _parent = self.filter_tree.add_item(self.filter_tree.invisibleRootItem(),[_filter_category["name"]],{'object':None},SWTW_GetIcon("60955eb2eb706d797c86bb61f88e23465040d406"),checkable=False)

            for _filter in _filter_category["filters"]:

                _checked = Qt.Unchecked

                if _filter.id in [_sel_filter.id for _sel_filter in self.selected_filters]:

                    _checked = Qt.Checked

                _tree_item = self.filter_tree.add_item(_parent,[_filter.name],{'object':_filter},SWTW_GetIcon("c505863c1711dbb978c424ee81d5f8674fa7ef9c"),checkable=True,checkstate=_checked)
                self.filter_tree.currentItemChanged.connect(self.tree_select)
                self.filter_tree.itemClicked.connect(self.item_clicked)

    def load(self):

        _error = False
        
        _stream = open(self.path, 'r')

        try:
            _data = yaml.load(_stream)
        except:
            _error = True

        if not _error:

            if self.name in _data.keys():

                if self.name == "Version":

                    self.version = _data[self.name]
                    
                else:

                    for _filter_category in _data[self.name]:

                        _filter_category_list = {"name":_filter_category,"filters":[]}

                        for _filter_name in _data[self.name][_filter_category]:

                            _filter = PTCC_Filter(_filter_name,_data[self.name][_filter_category][_filter_name])

                            _filter_category_list["filters"].append(_filter)

                        self.filters.append(_filter_category_list)     

        return _error

    def draw_gui(self):

        self.setWindowTitle("PTCC Filter %s" % (self.name,))
        self.setWindowIcon(SWTW_GetIcon('5a5160191638c3bd554f883f6d94b8e25ffc5e61'))
        self.setStyleSheet("background-color: #FFFFFF")
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)        

        self.draw_filter_tree()
        self.draw_filter_description()
        self.draw_filter_code()
        self.draw_operators()

        self.ly_left  = QVBoxLayout()
        self.ly_left.addLayout(self.ly_operators)
        self.ly_left.addWidget(self.filter_tree)
        

        self.ly_right = QVBoxLayout()
        self.ly_right.addWidget(self.filter_description)
        self.ly_right.addWidget(self.filter_code_bt)
        self.ly_right.addWidget(self.filter_code)
        self.ly_right.addWidget(self.filter_final_code)

        self.main_layout = QHBoxLayout()
        self.main_layout.addLayout(self.ly_left)
        self.main_layout.addLayout(self.ly_right)

        self.setLayout(self.main_layout)

        self.setFixedSize(800,600)

    def draw_operators(self):

        self.ly_operators = QHBoxLayout()
        self.filter_and = PTCC_WDG_RadioButton("AND")
        self.filter_or  = PTCC_WDG_RadioButton("OR")
        self.filter_or.setChecked(True)
        self.ly_operators.addWidget(self.filter_and)
        self.ly_operators.addWidget(self.filter_or)
        self.ly_operators.addItem(QSpacerItem(250,30))

        self.filter_and.pressed.connect(self.change_operator_and)
        self.filter_or.pressed.connect(self.change_operator_or)

    def draw_filter_tree(self):

        _CONTEXT_MENU_MODEL = [
                        {"Deselect all FIlters":{
                                        "icon":"2e153c9392586809e0cfc9fb5a7dcc5d2cf20b95",
                                        "callback":self.on_menu_action_deselect
                                        }
                        },
                      ]


        self.filter_tree = PTCC_WDG_Tree(usefind=True)
        self.filter_tree.setAnimated(True)
        self.filter_tree.header().setSectionResizeMode(0,QHeaderView.ResizeToContents)
        self.filter_tree.setFixedWidth(350)
        self.filter_tree.create_menu(_CONTEXT_MENU_MODEL)

    def draw_filter_description(self):

        self.filter_description = QTextEdit()
        self.filter_description.setReadOnly(True)
        self.filter_description.setText("")

    def draw_filter_code(self):

        self.filter_code_bt = PTCC_WDG_Small_Button(SWTW_GetIcon('4bac6ffa0dd45d560f48086df06c6a8880aef533'),SWTW_GetIcon('4bac6ffa0dd45d560f48086df06c6a8880aef533'),self.cbk_btn_code, "Show FIlter Code")
        self.filter_code    = QTextEdit()
        self.filter_code.setReadOnly(True)
        self.filter_code.setText("")
        self.filter_code.setVisible(self.code_visible)

        self.filter_final_code    = QTextEdit()
        self.filter_final_code.setReadOnly(True)
        self.filter_final_code.setText("")
        self.filter_final_code.setVisible(self.code_visible)

    def cbk_btn_code(self):

        self.code_visible = not self.code_visible
        self.filter_code.setVisible(self.code_visible)
        self.filter_final_code.setVisible(self.code_visible)

        if not self.code_visible:
            self.filter_code_bt.icon_normal = SWTW_GetIcon('4bac6ffa0dd45d560f48086df06c6a8880aef533')
            self.filter_code_bt.icon_hover  = SWTW_GetIcon('4bac6ffa0dd45d560f48086df06c6a8880aef533')
            self.filter_code_bt.setIcon(SWTW_GetIcon('4bac6ffa0dd45d560f48086df06c6a8880aef533'))
        else:
            self.filter_code_bt.icon_normal = SWTW_GetIcon('6efce44d9eeb810528ed14456cbe35e1a55491a6')
            self.filter_code_bt.icon_hover  = SWTW_GetIcon('6efce44d9eeb810528ed14456cbe35e1a55491a6')
            self.filter_code_bt.setIcon(SWTW_GetIcon('6efce44d9eeb810528ed14456cbe35e1a55491a6'))

    def item_clicked(self,item,column):

        pass

        _filter = self.filter_tree.get_user_data(item)["object"]

        if _filter != None:

            _status = item.checkState(0)

            if _status == Qt.Checked:
                if _filter.name not in [_sel_filter.name for _sel_filter in self.selected_filters]:
                    self.selected_filters.append(_filter)
            else:
                if _filter.name in [_sel_filter.name for _sel_filter in self.selected_filters]:
                    _index = [_sel_filter.name for _sel_filter in self.selected_filters].index(_filter.name)
                    del self.selected_filters[_index]

            self.code   = self.get_filter_code()
            self.filter = self.calculate_filter()

            self.filter_final_code.setText(self.code)

            self.parent.trigger_filter.emit(len(self.selected_filters) != 0)

    def tree_select(self, current, previous):

        if current != None:
            _user_data = self.filter_tree.get_user_data(current)
            _filter    = _user_data["object"]

            if _filter != None:           

                self.filter_description.setText(_filter.description)
                self.filter_code.setText(_filter.code)
                self.filter_final_code.setText(self.code)

    def change_operator_and(self):

        self.operator = "and"
        self.code = self.get_filter_code()
        self.filter_final_code.setText(self.code)

    def change_operator_or(self):

        self.operator = "or"
        self.code = self.get_filter_code()
        self.filter_final_code.setText(self.code)

    def get_filter_code(self):

        _code = ""

        for _index in range(len(self.selected_filters)):

            if len(self.selected_filters) == 1:
                _code += "%s" % (self.selected_filters[_index].code,)
            else:
                _code += "(%s)" % (self.selected_filters[_index].code,)

            if _index < (len(self.selected_filters) - 1):

                _code = "%s %s " % (_code,self.operator)

        return _code

    def calculate_filter(self):

        _filter = None

        if self.code != "":

            try:
                _filter = lambda item: eval(parser.expr(self.code).compile())  
            except:
                _filter = None

        return _filter

    def check(self,item):

        _state = True

        if self.filter != None:

            _state = self.filter(item)

        return _state

    def on_menu_action_deselect(self):

        for _index in range(self.filter_tree.topLevelItemCount()):
            
            _root = self.filter_tree.topLevelItem(_index)

            for _child_index in range(_root.childCount()):

                _filter = _root.child(_child_index)
                _filter.setCheckState(0,Qt.Unchecked)


        self.selected_filters = []
        self.code             = ""
        self.filter           = None 

        self.filter_final_code.setText(self.code)
        self.parent.trigger_filter.emit(len(self.selected_filters) != 0)

"""*************************************************************************************************
****************************************************************************************************
*************************************************************************************************"""
class PTCC_WDG_Help(QMainWindow):

    def __init__(self,parent):

        QMainWindow.__init__(self,parent)

        self.draw_gui()

    def draw_gui(self):

        self.setWindowTitle("Help")
        self.setWindowIcon(SWTW_GetIcon('5a5160191638c3bd554f883f6d94b8e25ffc5e61'))
        self.resize(400,300)
        self.setStyleSheet("background-color: #FFFFFF")
        self.setFixedSize(400,300)
        
        self.ctrl_wdg = QWidget()

        self.html = QTextBrowser()

        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.html)

        self.ctrl_wdg.setLayout(self.main_layout)

        self.setCentralWidget(self.ctrl_wdg)

        self.html.setOpenExternalLinks(True)

        self.clear_text()

    def set_text(self, help_id):

        _txt_html = PTCC_HELP(help_id)

        self.resize(_txt_html.width,_txt_html.height)
        self.setFixedSize(_txt_html.width,_txt_html.height)
        
        self.html.setHtml(_txt_html.text);

    def clear_text(self):
        _txt = """
        <p>&nbsp;</p>
<p>&nbsp;</p>
<p>&nbsp;</p>
<p><strong>Move Mouse over Buttons, Tables, Lists, Plots, Pie Charts to view information about them.</strong></p>
<p>&nbsp;</p>
        """
        self.html.setHtml(_txt);
